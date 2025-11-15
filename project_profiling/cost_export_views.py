# ========================================
# WEEKLY COST REPORT EXPORT VIEWS
# PDF and Excel Export Functionality
# ========================================

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.db.models import Sum
from datetime import datetime
import json
import base64
from io import BytesIO
from decimal import Decimal

from authentication.utils.decorators import verified_email_required, role_required
from .models import ProjectProfile, WeeklyCostReport
from .cost_tracking_views import aggregate_monthly_data, calculate_totals

# WeasyPrint for PDF generation
from weasyprint import HTML, CSS
import weasyprint

# OpenPyXL for Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

import sys
import os
import traceback as tb
import django


# ========================================
# DEBUG HELPER FOR PDF ERRORS
# ========================================

def render_pdf_error_debug(request, project_id, error, error_stage):
    """
    Render a detailed debug page for PDF generation errors
    """
    import traceback

    # Collect system information
    context = {
        'error_message': str(error),
        'error_type': type(error).__name__,
        'error_stage': error_stage,
        'project_id': project_id,
        'project_name': 'Unknown',
        'timestamp': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
        'traceback': traceback.format_exc(),
        'python_version': sys.version,
        'django_version': django.get_version(),
        'weasyprint_version': weasyprint.__version__ if hasattr(weasyprint, '__version__') else 'Unknown',
        'environment': os.environ.get('RENDER', 'Development'),
        'suggestions': []
    }

    # Get project name if available
    try:
        if project_id:
            project = ProjectProfile.objects.get(id=project_id)
            context['project_name'] = project.project_name
    except:
        pass

    # Check for Cairo/Pango installation
    try:
        import cairocffi
        context['cairo_installed'] = f'✓ Yes (version {cairocffi.version})'
    except ImportError as e:
        context['cairo_installed'] = f'✗ No - {str(e)}'
        context['suggestions'].append('Install Cairo library: Add "libcairo2" to apt.txt')

    try:
        import pangocffi
        context['pango_installed'] = '✓ Yes'
    except ImportError:
        context['pango_installed'] = '✗ No'
        context['suggestions'].append('Install Pango library: Add "libpango-1.0-0" to apt.txt')

    try:
        import gi
        gi.require_version('GdkPixbuf', '2.0')
        context['gdk_installed'] = '✓ Yes'
    except:
        context['gdk_installed'] = '✗ No'
        context['suggestions'].append('Install GDK-Pixbuf: Add "libgdk-pixbuf2.0-0" to apt.txt')

    # Check fonts
    try:
        font_dirs = ['/usr/share/fonts', '/usr/local/share/fonts']
        font_count = 0
        for font_dir in font_dirs:
            if os.path.exists(font_dir):
                for root, dirs, files in os.walk(font_dir):
                    font_count += len([f for f in files if f.endswith(('.ttf', '.otf'))])
        context['fonts_count'] = f'{font_count} fonts found'
    except:
        context['fonts_count'] = 'Unable to check'
        context['suggestions'].append('Install fonts: Add "fonts-liberation" to apt.txt')

    # Add common error suggestions
    error_str = str(error).lower()
    if 'cairo' in error_str:
        context['suggestions'].append('Cairo library error detected. Ensure all Cairo dependencies are installed.')
    if 'pango' in error_str:
        context['suggestions'].append('Pango library error detected. Ensure Pango and PangoCairo are installed.')
    if 'font' in error_str:
        context['suggestions'].append('Font-related error. Install system fonts or embed fonts in CSS.')
    if 'permission' in error_str or 'denied' in error_str:
        context['suggestions'].append('Permission error. Check file system permissions on Render.')

    # Create raw debug data for copying
    context['raw_debug_data'] = f"""
PDF Generation Error Debug Report
==================================
Timestamp: {context['timestamp']}
Project ID: {context['project_id']}
Project Name: {context['project_name']}
Error Stage: {error_stage}

Error Type: {context['error_type']}
Error Message: {context['error_message']}

System Information:
- Python: {context['python_version']}
- Django: {context['django_version']}
- WeasyPrint: {context['weasyprint_version']}
- Environment: {context['environment']}

Dependencies:
- Cairo: {context['cairo_installed']}
- Pango: {context['pango_installed']}
- GDK-Pixbuf: {context['gdk_installed']}
- Fonts: {context['fonts_count']}

Full Traceback:
{context['traceback']}
    """.strip()

    return render(request, 'debug/pdf_error.html', context)


# ========================================
# PDF EXPORT
# ========================================

@login_required
@verified_email_required
@role_required('EG', 'OM', 'PM')
def export_weekly_cost_pdf(request, project_id):
    """
    Export weekly cost reports to PDF with charts and tables
    """
    try:
        project = get_object_or_404(ProjectProfile, id=project_id)
        user_profile = request.user.userprofile

        # Check permissions
        if user_profile.role == 'PM' and project.project_manager != user_profile:
            return JsonResponse({'error': 'Unauthorized'}, status=403)

        # Get query params for date filtering
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        # Get POST body for include options and chart data
        if request.method == 'POST':
            body = json.loads(request.body)
            include_options = body.get('includeOptions', {})
            chart_data = body.get('chartData', {})
        else:
            include_options = {}
            chart_data = {}

        # Query weekly reports
        reports = WeeklyCostReport.objects.filter(project=project)
        if start_date and end_date:
            reports = reports.filter(
                period_start__gte=start_date,
                period_end__lte=end_date
            )

        # Aggregate data
        monthly_summary = aggregate_monthly_data(reports)
        totals = calculate_totals(reports)

        # Get logo path
        import os
        from django.conf import settings
        logo_path = os.path.join(settings.STATIC_ROOT or settings.BASE_DIR, 'powermason_capstone', 'static', 'img', 'powermason_logo.png')

        # Fallback to static directory if STATIC_ROOT doesn't have the file
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.BASE_DIR, 'powermason_capstone', 'static', 'img', 'powermason_logo.png')

        # Convert to file URL for WeasyPrint
        logo_url = f'file:///{logo_path.replace(os.sep, "/")}' if os.path.exists(logo_path) else None

        # Get Philippines timezone
        import pytz
        ph_tz = pytz.timezone('Asia/Manila')
        ph_time = timezone.now().astimezone(ph_tz)

        # Prepare context for PDF template
        context = {
            'project': project,
            'start_date': start_date or 'All',
            'end_date': end_date or 'All',
            'generated_date': ph_time.strftime('%B %d, %Y %I:%M %p') + ' (PHT)',
            'generated_by': request.user.get_full_name() or request.user.username,
            'weekly_reports': reports,
            'monthly_summary': monthly_summary,
            'totals': totals,
            'logo_path': logo_url,
            'include_project_header': include_options.get('project_header', True),
            'include_weekly_table': include_options.get('weekly_table', True),
            'include_monthly_table': include_options.get('monthly_table', True),
            'include_weekly_chart': include_options.get('weekly_chart', True),
            'include_monthly_chart': include_options.get('monthly_chart', True),
            'include_totals': include_options.get('totals', True),
            'include_category_breakdown': include_options.get('category_breakdown', True),
            'weekly_chart_image': chart_data.get('weeklyChart'),
            'monthly_chart_image': chart_data.get('monthlyChart'),
        }

        # Render HTML template
        html_string = render_to_string(
            'project_profiling/reports/weekly_cost_report_pdf.html',
            context
        )

        # Generate PDF
        try:
            html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
            pdf = html.write_pdf()

            # Validate PDF content
            if not pdf or len(pdf) < 100:
                raise ValueError(f"PDF generation produced invalid or empty file (size: {len(pdf) if pdf else 0} bytes)")

        except Exception as pdf_error:
            # Show debug page for PDF generation errors
            return render_pdf_error_debug(request, project_id, pdf_error, 'PDF Generation')

        # Create response
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f"{project.project_name.replace(' ', '_')}_Cost_Report_{start_date or 'All'}_to_{end_date or 'All'}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        # Show debug page for general errors
        return render_pdf_error_debug(request, project_id if 'project_id' in locals() else None, e, 'General')


# ========================================
# EXCEL EXPORT
# ========================================

@login_required
@verified_email_required
@role_required('EG', 'OM', 'PM')
def export_weekly_cost_excel(request, project_id):
    """
    Export weekly cost reports to Excel with multiple sheets
    """
    try:
        project = get_object_or_404(ProjectProfile, id=project_id)
        user_profile = request.user.userprofile

        # Check permissions
        if user_profile.role == 'PM' and project.project_manager != user_profile:
            return JsonResponse({'error': 'Unauthorized'}, status=403)

        # Get query params
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        # Get POST body
        if request.method == 'POST':
            body = json.loads(request.body)
            include_options = body.get('includeOptions', {})
            chart_data = body.get('chartData', {})
        else:
            include_options = {}
            chart_data = {}

        # Query data
        reports = WeeklyCostReport.objects.filter(project=project)
        if start_date and end_date:
            reports = reports.filter(
                period_start__gte=start_date,
                period_end__lte=end_date
            )

        monthly_summary = aggregate_monthly_data(reports)
        totals = calculate_totals(reports)

        # Create workbook
        wb = Workbook()

        # Sheet 1: Overview
        ws_overview = wb.active
        ws_overview.title = "Overview"
        create_overview_sheet(ws_overview, project, start_date, end_date, totals, include_options)

        # Sheet 2: Weekly Data
        if include_options.get('weekly_table', True):
            ws_weekly = wb.create_sheet("Weekly Data")
            create_weekly_sheet(ws_weekly, reports, chart_data.get('weeklyChart'), include_options)

        # Sheet 3: Monthly Data
        if include_options.get('monthly_table', True):
            ws_monthly = wb.create_sheet("Monthly Data")
            create_monthly_sheet(ws_monthly, monthly_summary, chart_data.get('monthlyChart'), include_options)

        # Sheet 4: Category Analysis
        if include_options.get('category_breakdown', True):
            ws_category = wb.create_sheet("Category Analysis")
            create_category_sheet(ws_category, reports, totals)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{project.project_name.replace(' ', '_')}_Cost_Report_{start_date or 'All'}_to_{end_date or 'All'}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ========================================
# EXCEL HELPER FUNCTIONS
# ========================================

def create_overview_sheet(ws, project, start_date, end_date, totals, include_options):
    """Create overview sheet with project details and totals"""
    # Title
    ws['A1'] = 'Weekly Cost Report - Overview'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')

    # Project details
    ws['A3'] = 'Project Name:'
    ws['B3'] = project.project_name
    ws['A4'] = 'Project Code:'
    ws['B4'] = project.project_id
    ws['A5'] = 'Report Period:'
    ws['B5'] = f"{start_date or 'All'} to {end_date or 'All'}"
    ws['A6'] = 'Generated:'

    # Get Philippines timezone for generated timestamp
    import pytz
    ph_tz = pytz.timezone('Asia/Manila')
    ph_time = timezone.now().astimezone(ph_tz)
    ws['B6'] = ph_time.strftime('%B %d, %Y %I:%M %p') + ' (PHT)'

    # Make labels bold
    for row in range(3, 7):
        ws[f'A{row}'].font = Font(bold=True)

    # Totals summary
    if include_options.get('totals', True):
        ws['A9'] = 'TOTAL DISBURSEMENT SUMMARY'
        ws['A9'].font = Font(size=14, bold=True)
        ws.merge_cells('A9:B9')

        headers = ['Category', 'Amount (₱)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(10, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        categories = [
            ('GENREQ', totals['genreq']),
            ('MATERIALS', totals['materials']),
            ('LABOR', totals['labor']),
            ('EQUIPMENT', totals['equipment']),
            ('TOTAL', totals['total'])
        ]

        for row, (category, amount) in enumerate(categories, 11):
            ws.cell(row, 1, category)
            ws.cell(row, 2, amount)
            ws.cell(row, 2).number_format = '#,##0.00'
            if category == 'TOTAL':
                ws.cell(row, 1).font = Font(bold=True)
                ws.cell(row, 2).font = Font(bold=True)
                ws.cell(row, 1).fill = PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid')
                ws.cell(row, 2).fill = PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid')

    # Auto-fit columns
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30


def create_weekly_sheet(ws, reports, chart_image, include_options):
    """Create weekly data sheet"""
    # Title
    ws['A1'] = 'Week-to-Week Summary'
    ws['A1'].font = Font(size=14, bold=True)

    # Headers
    headers = ['DATE', 'PERIOD', 'GENREQ', 'MATERIALS', 'LABOR', 'EQUIPMENT', 'TOTAL']
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    header_font = Font(bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    row = 4
    for report in reports:
        ws.cell(row, 1, report.report_date.strftime('%Y-%m-%d'))
        ws.cell(row, 2, f"{report.period_start.strftime('%Y-%m-%d')} - {report.period_end.strftime('%Y-%m-%d')}")
        ws.cell(row, 3, float(report.genreq_amount))
        ws.cell(row, 4, float(report.materials_amount))
        ws.cell(row, 5, float(report.labor_amount))
        ws.cell(row, 6, float(report.equipment_amount))
        ws.cell(row, 7, float(report.total_amount))
        row += 1

    # Totals row
    if include_options.get('totals', True):
        totals_fill = PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid')
        totals_font = Font(bold=True)

        ws.cell(row, 1, 'TOTAL DISBURSEMENT')
        ws.cell(row, 1).font = totals_font
        ws.cell(row, 1).fill = totals_fill
        ws.cell(row, 2, '')
        ws.cell(row, 2).fill = totals_fill

        for col in range(3, 8):
            ws.cell(row, col, f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{row-1})')
            ws.cell(row, col).font = totals_font
            ws.cell(row, col).fill = totals_fill

    # Format currency columns
    for r in range(4, row + 1):
        for col in range(3, 8):
            ws.cell(r, col).number_format = '#,##0.00'

    # Auto-fit columns
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Insert chart image if provided
    if chart_image:
        print(f"Weekly chart image received: {len(chart_image) if chart_image else 0} characters")
        if include_options.get('weekly_chart', True):
            try:
                # Add chart title
                ws.cell(row + 2, 1, 'Weekly Disbursement Chart')
                ws.cell(row + 2, 1).font = Font(size=12, bold=True)
                ws.merge_cells(f'A{row + 2}:G{row + 2}')
                insert_chart_image(ws, chart_image, row + 3)
                print(f"Weekly chart inserted successfully at row {row + 3}")
            except Exception as e:
                print(f"Failed to insert weekly chart: {str(e)}")
                import traceback
                traceback.print_exc()
    else:
        print("No weekly chart image provided")


def create_monthly_sheet(ws, monthly_summary, chart_image, include_options):
    """Create monthly data sheet"""
    # Title
    ws['A1'] = 'Month-to-Month Summary'
    ws['A1'].font = Font(size=14, bold=True)

    # Headers
    headers = ['MONTH', 'GENREQ', 'MATERIALS', 'LABOR', 'EQUIPMENT', 'TOTAL']
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    header_font = Font(bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    row = 4
    for month_data in monthly_summary:
        ws.cell(row, 1, month_data['month'])
        ws.cell(row, 2, month_data['genreq'])
        ws.cell(row, 3, month_data['materials'])
        ws.cell(row, 4, month_data['labor'])
        ws.cell(row, 5, month_data['equipment'])
        ws.cell(row, 6, month_data['total'])
        row += 1

    # Totals row
    if include_options.get('totals', True):
        totals_fill = PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid')
        totals_font = Font(bold=True)

        ws.cell(row, 1, 'TOTAL DISBURSEMENT')
        ws.cell(row, 1).font = totals_font
        ws.cell(row, 1).fill = totals_fill

        for col in range(2, 7):
            ws.cell(row, col, f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{row-1})')
            ws.cell(row, col).font = totals_font
            ws.cell(row, col).fill = totals_fill

    # Format currency columns
    for r in range(4, row + 1):
        for col in range(2, 7):
            ws.cell(r, col).number_format = '#,##0.00'

    # Auto-fit columns
    ws.column_dimensions['A'].width = 20
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Insert chart image if provided
    if chart_image:
        print(f"Monthly chart image received: {len(chart_image) if chart_image else 0} characters")
        if include_options.get('monthly_chart', True):
            try:
                # Add chart title
                ws.cell(row + 2, 1, 'Monthly Disbursement Chart')
                ws.cell(row + 2, 1).font = Font(size=12, bold=True)
                ws.merge_cells(f'A{row + 2}:F{row + 2}')
                insert_chart_image(ws, chart_image, row + 3)
                print(f"Monthly chart inserted successfully at row {row + 3}")
            except Exception as e:
                print(f"Failed to insert monthly chart: {str(e)}")
                import traceback
                traceback.print_exc()
    else:
        print("No monthly chart image provided")


def create_category_sheet(ws, reports, totals):
    """Create category breakdown analysis sheet"""
    # Title
    ws['A1'] = 'Category Breakdown Analysis'
    ws['A1'].font = Font(size=14, bold=True)

    # Category totals
    ws['A3'] = 'Category'
    ws['B3'] = 'Total Amount (₱)'
    ws['C3'] = 'Percentage'

    for col in range(1, 4):
        ws.cell(3, col).font = Font(bold=True)
        ws.cell(3, col).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    categories = [
        ('GENREQ', totals['genreq']),
        ('MATERIALS', totals['materials']),
        ('LABOR', totals['labor']),
        ('EQUIPMENT', totals['equipment'])
    ]

    total_amount = totals['total']
    row = 4
    for category, amount in categories:
        ws.cell(row, 1, category)
        ws.cell(row, 2, amount)
        ws.cell(row, 2).number_format = '#,##0.00'
        percentage = (amount / total_amount * 100) if total_amount > 0 else 0
        ws.cell(row, 3, percentage)
        ws.cell(row, 3).number_format = '0.00%'
        ws.cell(row, 3).value = percentage / 100  # Excel expects decimal for percentage
        row += 1

    # Auto-fit columns
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15


def insert_chart_image(ws, base64_image, start_row):
    """Insert chart image into Excel worksheet"""
    try:
        if not base64_image:
            return

        # Decode base64 image - handle both with and without data URL prefix
        if ',' in base64_image:
            image_data = base64.b64decode(base64_image.split(',')[1])
        else:
            image_data = base64.b64decode(base64_image)

        image = PILImage.open(BytesIO(image_data))

        # Convert to RGB if image has transparency (RGBA)
        if image.mode in ('RGBA', 'LA', 'P'):
            background = PILImage.new('RGB', image.size, (255, 255, 255))
            if image.mode == 'P':
                image = image.convert('RGBA')
            background.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
            image = background

        # Save to temp BytesIO
        img_io = BytesIO()
        image.save(img_io, format='PNG')
        img_io.seek(0)

        # Insert into Excel with proper sizing
        img = XLImage(img_io)
        # Maintain aspect ratio
        img.width = 700
        img.height = 450
        ws.add_image(img, f'A{start_row}')

    except Exception as e:
        # Log the error but continue execution
        print(f"Error inserting chart image: {str(e)}")
        pass


# ========================================
# DASHBOARD BUDGET SUMMARY API
# ========================================

@login_required
@verified_email_required
def api_dashboard_budget_summary(request):
    """
    API endpoint for dashboard budget chart
    Returns approved budget vs actual disbursement for all projects
    """
    try:
        user_profile = request.user.userprofile

        # Get projects based on role
        if user_profile.role in ['EG', 'OM']:
            projects = ProjectProfile.objects.all()
        elif user_profile.role == 'PM':
            projects = ProjectProfile.objects.filter(project_manager=user_profile)
        else:
            projects = ProjectProfile.objects.none()

        # Prepare data
        project_names = []
        approved_budgets = []
        actual_disbursements = []

        for project in projects:
            project_names.append(project.project_name)

            # Get approved budget (sum of all budget categories)
            from .models import ProjectBudget
            approved_budget = ProjectBudget.objects.filter(
                project_scope__project=project
            ).aggregate(
                total=Sum('planned_amount')
            )['total'] or 0
            approved_budgets.append(float(approved_budget))

            # Get actual disbursement (sum of all weekly reports)
            actual_disbursement = WeeklyCostReport.objects.filter(
                project=project
            ).aggregate(
                total=Sum('total_amount')
            )['total'] or 0
            actual_disbursements.append(float(actual_disbursement))

        return JsonResponse({
            'success': True,
            'project_names': project_names,
            'approved_budgets': approved_budgets,
            'actual_disbursements': actual_disbursements
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
