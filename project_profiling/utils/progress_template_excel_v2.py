"""
Progress Template Excel Generator V2
Creates Excel template matching user's format with weekly columns
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta


class ProgressTemplateExcelGeneratorV2:
    """Generate Excel template matching user's format"""

    def __init__(self, template_data):
        """
        Initialize with template data from progress_template_generator

        Args:
            template_data: Dict with project, divisions, BOQ items
        """
        self.template_data = template_data
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Progress Report"

        # Parse week dates
        self.week_start = datetime.fromisoformat(template_data['week_start_date'])
        self.week_end = datetime.fromisoformat(template_data['week_end_date'])

        # Styles
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')

        self.division_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        self.division_font = Font(name='Arial', size=9, bold=True)

        self.subtotal_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.subtotal_font = Font(name='Arial', size=9, bold=True)

        self.editable_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.readonly_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.current_row = 1

    def generate_template(self):
        """Generate Excel template"""
        self._add_header_section()
        self._add_column_headers()
        self._add_boq_items()
        self._add_grand_total()
        self._format_columns()
        self._freeze_panes()

        return self.wb

    def _add_header_section(self):
        """Add project info header matching user's format"""
        # Row 1: PROJ ID and PERIOD
        self.ws['A1'] = 'PROJ ID'
        self.ws['A1'].font = Font(name='Arial', size=9, bold=True)
        # Use project_code (like GC-057) instead of numeric ID
        self.ws['B1'] = self.template_data.get('project_code', self.template_data['project_id'])
        self.ws['B1'].font = Font(name='Arial', size=9)

        # Calculate last date column for PERIOD placement
        num_days = (self.week_end - self.week_start).days + 1
        period_col = get_column_letter(6 + num_days)  # After all daily columns

        self.ws[f'{period_col}1'] = f"PERIOD: {self.week_start.strftime('%b %d')} - {self.week_end.strftime('%d, %Y')}"
        self.ws[f'{period_col}1'].font = Font(name='Arial', size=9, bold=True)
        self.ws[f'{period_col}1'].alignment = Alignment(horizontal='right')

        # Row 2: PROJECT and AS OF
        self.ws['A2'] = 'PROJECT'
        self.ws['A2'].font = Font(name='Arial', size=9, bold=True)
        self.ws.merge_cells('B2:D2')
        self.ws['B2'] = self.template_data['project_name']
        self.ws['B2'].font = Font(name='Arial', size=9)

        self.ws[f'{period_col}2'] = f"AS OF {datetime.now().strftime('%b %d, %Y')}"
        self.ws[f'{period_col}2'].font = Font(name='Arial', size=9)
        self.ws[f'{period_col}2'].alignment = Alignment(horizontal='right')

        # Row 3: STARTED and PROGRESS THIS WEEK indicator
        self.ws['A3'] = 'STARTED'
        self.ws['A3'].font = Font(name='Arial', size=9, bold=True)
        self.ws['B3'] = self.week_start.strftime('%B %d, %Y')
        self.ws['B3'].font = Font(name='Arial', size=9)

        # Merge cells for PROGRESS THIS WEEK indicator
        progress_col_start = get_column_letter(6)  # First daily column
        progress_col_end = get_column_letter(6 + num_days - 1)  # Last daily column
        self.ws.merge_cells(f'{progress_col_start}3:{progress_col_end}3')
        self.ws[f'{progress_col_start}3'] = 'PROGRESS THIS WEEK'
        self.ws[f'{progress_col_start}3'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.ws[f'{progress_col_start}3'].font = Font(name='Arial', size=10, bold=True)
        self.ws[f'{progress_col_start}3'].alignment = Alignment(horizontal='center', vertical='center')

        # Add percentage in right cell (will be auto-calculated from totals)
        percent_cell = get_column_letter(6 + num_days)
        # Formula: Sum of all period amounts / Total approved amount
        amount_col = get_column_letter(6 + num_days)  # TOTAL AMOUNT column
        self.ws[f'{percent_cell}3'] = ''  # Leave blank - will show formula result later
        self.ws[f'{percent_cell}3'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.ws[f'{percent_cell}3'].font = Font(name='Arial', size=10, bold=True)
        self.ws[f'{percent_cell}3'].alignment = Alignment(horizontal='right', vertical='center')
        self.ws[f'{percent_cell}3'].number_format = '0.00%'

        self.current_row = 5

    def _add_column_headers(self):
        """Add column headers matching user's format"""
        header_row1 = self.current_row
        header_row2 = self.current_row + 1

        # Column A: ITEM (merged across 2 rows)
        self.ws.merge_cells(f'A{header_row1}:A{header_row2}')
        cell = self.ws[f'A{header_row1}']
        cell.value = 'ITEM'
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.thin_border

        # Column B: PARTICULARS (merged across 2 rows)
        self.ws.merge_cells(f'B{header_row1}:B{header_row2}')
        cell = self.ws[f'B{header_row1}']
        cell.value = 'PARTICULARS'
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.thin_border

        # Columns C-E: APPROVED CONTRACT (merged in row 1)
        self.ws.merge_cells(f'C{header_row1}:E{header_row1}')
        cell = self.ws[f'C{header_row1}']
        cell.value = 'APPROVED CONTRACT'
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.thin_border

        # Row 2 sub-headers
        self.ws[f'C{header_row2}'] = 'QTY'
        self.ws[f'C{header_row2}'].fill = self.header_fill
        self.ws[f'C{header_row2}'].font = self.header_font
        self.ws[f'C{header_row2}'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws[f'C{header_row2}'].border = self.thin_border

        self.ws[f'D{header_row2}'] = 'UNIT'
        self.ws[f'D{header_row2}'].fill = self.header_fill
        self.ws[f'D{header_row2}'].font = self.header_font
        self.ws[f'D{header_row2}'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws[f'D{header_row2}'].border = self.thin_border

        self.ws[f'E{header_row2}'] = 'AMOUNT'
        self.ws[f'E{header_row2}'].fill = self.header_fill
        self.ws[f'E{header_row2}'].font = self.header_font
        self.ws[f'E{header_row2}'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws[f'E{header_row2}'].border = self.thin_border

        # Add daily date columns
        current_date = self.week_start
        date_col_idx = 6
        date_cols = []

        while current_date <= self.week_end:
            # Format: "12-Aug-24" in row 1, "MON" in row 2
            date_str = current_date.strftime('%d-%b-%y')
            day_str = current_date.strftime('%a').upper()

            # Row 1: Date
            cell = self.ws.cell(row=header_row1, column=date_col_idx)
            cell.value = date_str
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border

            # Row 2: Day
            cell = self.ws.cell(row=header_row2, column=date_col_idx)
            cell.value = day_str
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border

            date_cols.append(date_col_idx)
            date_col_idx += 1
            current_date += timedelta(days=1)

        # TOTAL section (merged in row 1)
        total_col_start = date_col_idx
        total_col_end = date_col_idx + 1

        self.ws.merge_cells(
            start_row=header_row1, start_column=total_col_start,
            end_row=header_row1, end_column=total_col_end
        )
        cell = self.ws.cell(row=header_row1, column=total_col_start)
        cell.value = 'TOTAL'
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.thin_border

        # Row 2: AMOUNT and PERCENT
        cell = self.ws.cell(row=header_row2, column=total_col_start)
        cell.value = 'AMOUNT'
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.thin_border

        cell = self.ws.cell(row=header_row2, column=total_col_end)
        cell.value = 'PERCENT'
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.thin_border

        self.current_row = header_row2 + 1
        self.data_start_row = self.current_row
        self.num_date_cols = len(date_cols)

    def _add_boq_items(self):
        """Add BOQ items grouped by division"""
        if not self.template_data.get('divisions'):
            return

        for div_idx, division in enumerate(self.template_data['divisions'], start=1):
            # Division header - format: "DIV 1    GENERAL REQUIREMENTS"
            cell_a = self.ws.cell(row=self.current_row, column=1)
            cell_a.value = f"DIV {div_idx}"
            cell_a.font = self.division_font
            cell_a.fill = self.division_fill
            cell_a.border = self.thin_border
            cell_a.alignment = Alignment(horizontal='center', vertical='center')

            cell_b = self.ws.cell(row=self.current_row, column=2)
            cell_b.value = division['name']
            cell_b.font = self.division_font
            cell_b.fill = self.division_fill
            cell_b.border = self.thin_border
            cell_b.alignment = Alignment(horizontal='left', vertical='center')

            # Apply styling to all cells in the division header row
            last_col = 5 + self.num_date_cols + 2
            for col in range(3, last_col + 1):
                cell = self.ws.cell(row=self.current_row, column=col)
                cell.fill = self.division_fill
                cell.border = self.thin_border

            self.current_row += 1

            # Add BOQ items
            for item in division['boq_items']:
                self._add_boq_item_row(item)

            # Division subtotal
            self._add_division_subtotal(division['name'])

    def _add_boq_item_row(self, item):
        """Add a single BOQ item row"""
        row = self.current_row

        # ITEM (BOQ Code)
        cell = self.ws.cell(row=row, column=1)
        cell.value = item['code']
        cell.fill = self.readonly_fill
        cell.border = self.thin_border
        cell.alignment = Alignment(horizontal='left')
        cell.font = Font(name='Arial', size=9)

        # PARTICULARS (Description)
        cell = self.ws.cell(row=row, column=2)
        cell.value = item['description']
        cell.fill = self.readonly_fill
        cell.border = self.thin_border
        cell.alignment = Alignment(horizontal='left', wrap_text=True)
        cell.font = Font(name='Arial', size=9)

        # QTY
        cell = self.ws.cell(row=row, column=3)
        cell.value = float(item.get('quantity', 0))
        cell.fill = self.readonly_fill
        cell.border = self.thin_border
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        cell.font = Font(name='Arial', size=9)

        # UNIT
        cell = self.ws.cell(row=row, column=4)
        cell.value = item.get('uom', '')
        cell.fill = self.readonly_fill
        cell.border = self.thin_border
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(name='Arial', size=9)

        # AMOUNT
        cell = self.ws.cell(row=row, column=5)
        cell.value = float(item['approved_amount'])
        cell.fill = self.readonly_fill
        cell.border = self.thin_border
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        cell.font = Font(name='Arial', size=9)

        # Daily columns (EDITABLE - Yellow)
        for col_offset in range(self.num_date_cols):
            col = 6 + col_offset
            cell = self.ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = self.editable_fill
            cell.border = self.thin_border
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            cell.font = Font(name='Arial', size=9)

        # TOTAL AMOUNT (formula)
        total_col = 6 + self.num_date_cols
        cell = self.ws.cell(row=row, column=total_col)
        first_day_col = get_column_letter(6)
        last_day_col = get_column_letter(6 + self.num_date_cols - 1)
        cell.value = f"=SUM({first_day_col}{row}:{last_day_col}{row})"
        cell.fill = self.readonly_fill
        cell.border = self.thin_border
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        cell.font = Font(name='Arial', size=9)

        # PERCENT (formula)
        percent_col = total_col + 1
        cell = self.ws.cell(row=row, column=percent_col)
        amount_col = get_column_letter(5)
        total_amount_col = get_column_letter(total_col)
        cell.value = f"=IF({amount_col}{row}=0,0,{total_amount_col}{row}/{amount_col}{row})"
        cell.fill = self.readonly_fill
        cell.border = self.thin_border
        cell.number_format = '0.00%'
        cell.alignment = Alignment(horizontal='right')
        cell.font = Font(name='Arial', size=9)

        self.current_row += 1

    def _add_division_subtotal(self, division_name):
        """Add division subtotal row"""
        row = self.current_row

        # Label
        cell = self.ws.cell(row=row, column=2)
        cell.value = f"SUB-TOTAL FOR {division_name}"
        cell.fill = self.subtotal_fill
        cell.font = self.subtotal_font
        cell.border = self.thin_border
        cell.alignment = Alignment(horizontal='left')

        # Subtotal amount
        total_col = 6 + self.num_date_cols
        cell = self.ws.cell(row=row, column=total_col)
        # Sum all items above (from data_start_row to current_row-1)
        cell.value = f"=SUBTOTAL(9,{get_column_letter(total_col)}{self.data_start_row}:{get_column_letter(total_col)}{row-1})"
        cell.fill = self.subtotal_fill
        cell.font = self.subtotal_font
        cell.border = self.thin_border
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')

        self.current_row += 1

    def _add_grand_total(self):
        """Add grand total row"""
        row = self.current_row + 1

        # TOTAL label in column A
        cell = self.ws.cell(row=row, column=1)
        cell.value = "TOTAL"
        cell.fill = self.subtotal_fill
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.border = self.thin_border
        cell.alignment = Alignment(horizontal='center')

        # Label in column B
        cell = self.ws.cell(row=row, column=2)
        cell.value = "APPROVED WORKS (INCLUSIVE)"
        cell.fill = self.subtotal_fill
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.border = self.thin_border

        # Total Approved Contract Amount (Column E/5)
        cell = self.ws.cell(row=row, column=5)
        cell.value = f"=SUBTOTAL(9,{get_column_letter(5)}:{get_column_letter(5)})"
        cell.fill = self.subtotal_fill
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.border = self.thin_border
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')

        # Grand total amount (Total weekly progress)
        total_col = 6 + self.num_date_cols
        cell = self.ws.cell(row=row, column=total_col)
        cell.value = f"=SUBTOTAL(9,{get_column_letter(total_col)}:{get_column_letter(total_col)})"
        cell.fill = self.subtotal_fill
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.border = self.thin_border
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')

        # Grand total percent
        percent_col = total_col + 1
        cell = self.ws.cell(row=row, column=percent_col)
        # Use the total from column E (row value) instead of SUM to reference the total we just calculated
        cell.value = f"=IF({get_column_letter(5)}{row}=0,0,{get_column_letter(total_col)}{row}/{get_column_letter(5)}{row})"
        cell.fill = self.subtotal_fill
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.border = self.thin_border
        cell.number_format = '0.00%'
        cell.alignment = Alignment(horizontal='right')

    def _format_columns(self):
        """Set column widths"""
        self.ws.column_dimensions['A'].width = 8   # ITEM
        self.ws.column_dimensions['B'].width = 50  # PARTICULARS
        self.ws.column_dimensions['C'].width = 8   # QTY
        self.ws.column_dimensions['D'].width = 8   # UNIT
        self.ws.column_dimensions['E'].width = 12  # AMOUNT

        # Daily columns
        for col_offset in range(self.num_date_cols):
            col_letter = get_column_letter(6 + col_offset)
            self.ws.column_dimensions[col_letter].width = 10

        # TOTAL columns
        total_col = get_column_letter(6 + self.num_date_cols)
        self.ws.column_dimensions[total_col].width = 12

        percent_col = get_column_letter(6 + self.num_date_cols + 1)
        self.ws.column_dimensions[percent_col].width = 10

    def _freeze_panes(self):
        """Freeze header rows and first columns"""
        self.ws.freeze_panes = 'C7'  # Freeze first 6 rows and columns A-B


def generate_blank_template_v2(template_data):
    """
    Generate blank Excel template in user's format

    Args:
        template_data: Dict from progress_template_generator

    Returns:
        Workbook: Excel workbook for download
    """
    generator = ProgressTemplateExcelGeneratorV2(template_data)
    return generator.generate_template()
