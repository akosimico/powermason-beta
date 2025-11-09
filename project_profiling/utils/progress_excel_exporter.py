"""
Progress Excel Exporter
Generates Excel reports for weekly progress in BOQ format
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from decimal import Decimal
import pytz


class ProgressExcelExporter:
    """Export weekly progress reports to Excel in BOQ format"""

    def __init__(self, report):
        """
        Initialize exporter with a WeeklyProgressReport instance

        Args:
            report: WeeklyProgressReport instance
        """
        self.report = report
        self.project = report.project
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = f"Week {report.week_start_date.strftime('%m-%d')}"

        # Define styles
        self.header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        self.division_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.division_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        self.task_font = Font(name='Arial', size=11, bold=True)
        self.task_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        self.boq_font = Font(name='Arial', size=10)
        self.bold_font = Font(name='Arial', size=10, bold=True)

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.current_row = 1

    def export_weekly_report(self):
        """
        Export weekly progress report to Excel

        Returns:
            Workbook: openpyxl Workbook instance
        """
        self._add_project_info_header()
        self._add_column_headers()
        self._add_boq_items()
        self._add_totals()
        self._format_columns()

        return self.wb

    def _add_project_info_header(self):
        """Add project information header at the top"""
        # Calculate the column for right-side info (PERIOD and AS OF)
        # Place it far enough to the right to avoid conflicts
        week_days = self._get_week_days()
        num_days = len(week_days)
        right_col = max(9, 6 + num_days)  # At least column I, or beyond the daily columns

        # PROJ ID and PERIOD on same row
        cell = self.ws.cell(row=1, column=1)
        cell.value = 'PROJ ID'
        cell.font = Font(name='Arial', size=11, bold=True)

        cell = self.ws.cell(row=1, column=2)
        cell.value = self.project.project_id or f'PROJ-{self.project.id}'
        cell.font = Font(name='Arial', size=11)

        # Period info (right side)
        period_text = f"PERIOD: {self.report.week_start_date.strftime('%b %d')} - {self.report.week_end_date.strftime('%d, %Y')}"
        cell = self.ws.cell(row=1, column=right_col)
        cell.value = period_text
        cell.font = Font(name='Arial', size=11, bold=True)

        # PROJECT and AS OF on next row
        cell = self.ws.cell(row=2, column=1)
        cell.value = 'PROJECT'
        cell.font = Font(name='Arial', size=11, bold=True)

        cell = self.ws.cell(row=2, column=2)
        cell.value = self.project.project_name
        cell.font = Font(name='Arial', size=11)

        # Get Philippine time for AS OF
        ph_tz = pytz.timezone('Asia/Manila')
        ph_time = datetime.now(ph_tz)
        as_of_text = f"AS OF {ph_time.strftime('%b %d, %Y')}"
        cell = self.ws.cell(row=2, column=right_col)
        cell.value = as_of_text
        cell.font = Font(name='Arial', size=11, bold=True)

        # STARTED date
        if hasattr(self.project, 'start_date') and self.project.start_date:
            cell = self.ws.cell(row=3, column=1)
            cell.value = 'STARTED'
            cell.font = Font(name='Arial', size=11, bold=True)

            cell = self.ws.cell(row=3, column=2)
            cell.value = self.project.start_date.strftime('%B %d, %Y')
            cell.font = Font(name='Arial', size=11)

        self.current_row = 5

    def _get_week_days(self):
        """Generate list of days in the week"""
        days = []
        current_date = self.report.week_start_date
        while current_date <= self.report.week_end_date:
            days.append(current_date)
            current_date += timedelta(days=1)
        return days

    def _add_column_headers(self):
        """Add column headers for BOQ items with daily breakdown"""
        week_days = self._get_week_days()
        day_names = ['SUN', 'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT']

        # Row 1: Main headers
        # ITEM (col 1)
        cell = self.ws.cell(row=self.current_row, column=1)
        cell.value = 'ITEM'
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = self.border

        # PARTICULARS (col 2)
        cell = self.ws.cell(row=self.current_row, column=2)
        cell.value = 'PARTICULARS'
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = self.border

        # APPROVED CONTRACT (cols 3-5, merged)
        cell = self.ws.cell(row=self.current_row, column=3)
        cell.value = 'APPROVED CONTRACT'
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = self.border
        self.ws.merge_cells(start_row=self.current_row, start_column=3, end_row=self.current_row, end_column=5)

        # Apply style to merged cells
        for col in range(3, 6):
            self.ws.cell(row=self.current_row, column=col).font = self.header_font
            self.ws.cell(row=self.current_row, column=col).fill = self.header_fill
            self.ws.cell(row=self.current_row, column=col).border = self.border

        # Daily columns (starting from col 6)
        for i, day in enumerate(week_days):
            col_num = 6 + i
            cell = self.ws.cell(row=self.current_row, column=col_num)
            cell.value = day.strftime('%d-%b-%y')
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        # TOTAL (last 2 columns, merged)
        total_start_col = 6 + len(week_days)
        cell = self.ws.cell(row=self.current_row, column=total_start_col)
        cell.value = 'TOTAL'
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = self.border
        self.ws.merge_cells(start_row=self.current_row, start_column=total_start_col, end_row=self.current_row, end_column=total_start_col + 1)

        # Apply style to merged TOTAL cells
        for col in range(total_start_col, total_start_col + 2):
            self.ws.cell(row=self.current_row, column=col).font = self.header_font
            self.ws.cell(row=self.current_row, column=col).fill = self.header_fill
            self.ws.cell(row=self.current_row, column=col).border = self.border

        self.current_row += 1

        # Row 2: Sub-headers
        sub_headers = ['', '', 'QTY', 'UNIT', 'AMOUNT']
        # Add day names
        for day in week_days:
            day_name = day_names[day.weekday()]
            sub_headers.append(day_name)
        # Add TOTAL sub-headers
        sub_headers.extend(['AMOUNT', 'PERCENT'])

        for col_num, header in enumerate(sub_headers, start=1):
            cell = self.ws.cell(row=self.current_row, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        self.current_row += 1

    def _add_boq_items(self):
        """Add BOQ items grouped by division with daily breakdown"""
        from project_profiling.models import BOQItemProgress

        # Get all BOQ items for this report
        boq_items = BOQItemProgress.objects.filter(
            weekly_report=self.report
        ).select_related('project_task').order_by('division', 'task_group', 'boq_item_code')

        week_days = self._get_week_days()
        num_days = len(week_days)

        # Group by division
        current_division = None
        division_subtotals = {}

        for item in boq_items:
            # Add division header if changed
            if current_division != item.division:
                # Add subtotal for previous division
                if current_division is not None:
                    self._add_division_subtotal(current_division, division_subtotals.get(current_division, 0), num_days)

                # Division header
                cell = self.ws.cell(row=self.current_row, column=1)
                cell.value = item.division
                cell.font = self.division_font
                cell.fill = self.division_fill
                cell.alignment = Alignment(horizontal='left', vertical='center')

                # Merge across all columns (dynamically based on number of days)
                total_columns = 5 + num_days + 2  # ITEM + PARTICULARS + 3 APPROVED + days + AMOUNT + PERCENT
                self.ws.merge_cells(
                    start_row=self.current_row,
                    start_column=1,
                    end_row=self.current_row,
                    end_column=total_columns
                )

                for col in range(1, total_columns + 1):
                    self.ws.cell(row=self.current_row, column=col).border = self.border
                    self.ws.cell(row=self.current_row, column=col).fill = self.division_fill

                self.current_row += 1
                current_division = item.division
                division_subtotals[current_division] = 0

            # BOQ item row data
            row_data = [
                item.boq_item_code,  # Col A
                item.description,    # Col B
                float(item.quantity) if item.quantity else 0,  # Col C
                item.unit_of_measurement,  # Col D
                float(item.approved_contract_amount),  # Col E
            ]

            # Add daily progress (distribute period amount across the week)
            # For now, show the amount only in the last column (simplified)
            period_amount = float(item.period_progress_amount)
            for i in range(num_days):
                if i == num_days - 1:  # Show amount on last day
                    row_data.append(period_amount)
                else:
                    row_data.append(0)  # Empty cells for other days

            # Add totals
            row_data.append(period_amount)  # Total Amount
            row_data.append(float(item.period_progress_percent))  # Total Percent

            # Track division subtotal
            division_subtotals[current_division] += period_amount

            # Write row
            for col_num, value in enumerate(row_data, start=1):
                cell = self.ws.cell(row=self.current_row, column=col_num)
                cell.value = value
                cell.font = self.boq_font
                cell.border = self.border

                # Format numbers
                if col_num == 3:  # Quantity
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_num == 5 or col_num >= 6 + num_days:  # Amounts (approved and totals)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_num >= 6 and col_num < 6 + num_days:  # Daily amounts
                    if value > 0:
                        cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_num == 6 + num_days + 1:  # Percent
                    cell.number_format = '0.00"%"'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:  # Text columns
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            # Highlight if progress decreased
            if item.progress_decreased:
                for col in range(1, len(row_data) + 1):
                    self.ws.cell(row=self.current_row, column=col).fill = PatternFill(
                        start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'
                    )

            self.current_row += 1

        # Add subtotal for last division
        if current_division is not None:
            self._add_division_subtotal(current_division, division_subtotals.get(current_division, 0), num_days)

    def _add_division_subtotal(self, division_name, subtotal_amount, num_days):
        """Add subtotal row for a division"""
        cell = self.ws.cell(row=self.current_row, column=1)
        cell.value = f"SUB-TOTAL FOR {division_name}"
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self.border

        # Merge first columns
        total_columns = 5 + num_days + 2
        self.ws.merge_cells(
            start_row=self.current_row,
            start_column=1,
            end_row=self.current_row,
            end_column=5 + num_days
        )

        for col in range(1, 5 + num_days + 1):
            self.ws.cell(row=self.current_row, column=col).border = self.border
            self.ws.cell(row=self.current_row, column=col).fill = PatternFill(
                start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'
            )

        # Subtotal amount
        amount_col = 6 + num_days
        cell = self.ws.cell(row=self.current_row, column=amount_col)
        cell.value = subtotal_amount
        cell.number_format = '#,##0.00'
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        cell.border = self.border
        cell.alignment = Alignment(horizontal='right', vertical='center')

        # Empty percent cell
        percent_col = amount_col + 1
        cell = self.ws.cell(row=self.current_row, column=percent_col)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        cell.border = self.border

        self.current_row += 1

    def _add_totals(self):
        """Add totals row"""
        week_days = self._get_week_days()
        num_days = len(week_days)

        # Totals label
        cell = self.ws.cell(row=self.current_row, column=1)
        cell.value = "TOTAL"
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
        cell.border = self.border
        cell.alignment = Alignment(horizontal='left', vertical='center')

        # Merge first columns (up to before daily columns)
        merge_end_col = 5 + num_days
        self.ws.merge_cells(
            start_row=self.current_row,
            start_column=1,
            end_row=self.current_row,
            end_column=merge_end_col
        )

        for col in range(1, merge_end_col + 1):
            self.ws.cell(row=self.current_row, column=col).border = self.border
            self.ws.cell(row=self.current_row, column=col).fill = PatternFill(
                start_color='D0CECE', end_color='D0CECE', fill_type='solid'
            )

        # Total amount
        amount_col = 6 + num_days
        cell = self.ws.cell(row=self.current_row, column=amount_col)
        cell.value = float(self.report.total_period_amount)
        cell.number_format = '#,##0.00'
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
        cell.border = self.border
        cell.alignment = Alignment(horizontal='right', vertical='center')

        # Total percent
        percent_col = amount_col + 1
        cell = self.ws.cell(row=self.current_row, column=percent_col)
        cell.value = float(self.report.total_period_percent)
        cell.number_format = '0.00"%"'
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
        cell.border = self.border
        cell.alignment = Alignment(horizontal='right', vertical='center')

        self.current_row += 1

    def _format_columns(self):
        """Set column widths based on the new format"""
        week_days = self._get_week_days()
        num_days = len(week_days)

        # Set widths for fixed columns
        self.ws.column_dimensions['A'].width = 10   # ITEM
        self.ws.column_dimensions['B'].width = 40   # PARTICULARS
        self.ws.column_dimensions['C'].width = 10   # QTY
        self.ws.column_dimensions['D'].width = 8    # UNIT
        self.ws.column_dimensions['E'].width = 15   # AMOUNT

        # Set widths for daily columns (cols 6 onwards)
        for i in range(num_days):
            col_num = 6 + i
            col_letter = get_column_letter(col_num)
            self.ws.column_dimensions[col_letter].width = 12

        # Set widths for total columns
        total_amount_col_num = 6 + num_days
        total_percent_col_num = total_amount_col_num + 1
        self.ws.column_dimensions[get_column_letter(total_amount_col_num)].width = 15
        self.ws.column_dimensions[get_column_letter(total_percent_col_num)].width = 12


def export_weekly_report_to_excel(report):
    """
    Export a weekly progress report to Excel

    Args:
        report: WeeklyProgressReport instance

    Returns:
        Workbook: openpyxl Workbook instance ready to save
    """
    exporter = ProgressExcelExporter(report)
    return exporter.export_weekly_report()


def export_multiple_reports_to_excel(reports):
    """
    Export multiple weekly progress reports to a single Excel file
    with each report in a separate sheet

    Args:
        reports: QuerySet or list of WeeklyProgressReport instances

    Returns:
        Workbook: openpyxl Workbook instance with multiple sheets
    """
    if not reports:
        # Return empty workbook with message
        wb = Workbook()
        ws = wb.active
        ws.title = "No Reports"
        ws['A1'] = "No reports to export"
        return wb

    # Create first report
    first_report = reports[0]
    exporter = ProgressExcelExporter(first_report)
    wb = exporter.export_weekly_report()

    # Add remaining reports as new sheets
    for report in reports[1:]:
        # Create new exporter for each report
        exporter = ProgressExcelExporter(report)
        exporter.wb = wb
        exporter.ws = wb.create_sheet(title=f"Week {report.week_start_date.strftime('%m-%d')}")
        exporter.export_weekly_report()

    return wb
