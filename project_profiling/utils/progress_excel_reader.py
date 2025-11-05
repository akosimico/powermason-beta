"""
Progress Excel Reader
Reads and validates uploaded Excel progress reports
"""

from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ProgressExcelReader:
    """Read and extract data from uploaded Excel progress reports"""

    def __init__(self, file_path_or_stream):
        """
        Initialize reader with Excel file

        Args:
            file_path_or_stream: File path or file-like object
        """
        self.wb = load_workbook(file_path_or_stream, data_only=True)
        self.ws = self.wb.active
        self.errors = []
        self.warnings = []
        self.boq_items = []
        self.summary = {
            'total_items': 0,
            'items_with_progress': 0,
            'total_period_percent': Decimal('0'),
            'total_period_amount': Decimal('0'),
        }

    def read_and_validate(self):
        """
        Read Excel file and extract BOQ progress data

        Returns:
            dict: {
                'success': bool,
                'boq_items': list,
                'summary': dict,
                'errors': list,
                'warnings': list
            }
        """
        try:
            self._find_data_start()
            self._extract_boq_items()
            self._calculate_summary()
            self._validate_data()

            return {
                'success': len(self.errors) == 0,
                'boq_items': self.boq_items,
                'summary': self.summary,
                'errors': self.errors,
                'warnings': self.warnings
            }

        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            return {
                'success': False,
                'boq_items': [],
                'summary': {},
                'errors': [f"Failed to read Excel file: {str(e)}"],
                'warnings': []
            }

    def _find_data_start(self):
        """Find where BOQ data starts (after headers)"""
        # Look for "ITEM" or "BOQ Code" header (support both V1 and V2 templates)
        self.data_start_row = None

        for row_idx in range(1, min(50, self.ws.max_row + 1)):
            cell_value = self.ws[f'A{row_idx}'].value
            if cell_value and ('ITEM' in str(cell_value).upper() or 'BOQ Code' in str(cell_value)):
                self.data_start_row = row_idx + 1
                break

        if not self.data_start_row:
            self.errors.append("Could not find BOQ data table in Excel file. Please use the downloaded template.")
            raise ValueError("Data table not found")

        # Extract weekly progress summary from row 3 (V2 template format)
        self._extract_weekly_summary()

    def _extract_weekly_summary(self):
        """Extract weekly progress summary from row 3 (PROGRESS THIS WEEK row)"""
        # In V2 template, row 3 has the weekly progress percentage in the last column
        # Scan row 3 to find the percentage value (rightmost non-empty cell)
        try:
            for col_idx in range(self.ws.max_column, 0, -1):
                cell = self.ws.cell(row=3, column=col_idx)
                value = cell.value

                if value is not None and value != '' and value != 'PROGRESS THIS WEEK':
                    # Found the percentage value
                    if isinstance(value, (int, float)):
                        # Already a number (percentage as decimal, e.g., 0.019 = 1.9%)
                        weekly_percent = Decimal(str(value * 100))  # Convert to percentage
                    elif isinstance(value, str) and '%' in value:
                        # String with % sign
                        weekly_percent = self._parse_decimal(value, 3, 'Weekly Progress %')
                    else:
                        weekly_percent = self._parse_decimal(value, 3, 'Weekly Progress %')

                    self.summary['weekly_progress_percent'] = weekly_percent
                    logger.info(f"Found weekly progress in row 3: {weekly_percent}%")
                    break
        except Exception as e:
            logger.warning(f"Could not extract weekly summary from row 3: {str(e)}")
            # Not critical, continue anyway

    def _extract_boq_items(self):
        """Extract BOQ items from Excel"""
        current_division = None

        # Find AMOUNT and PERCENT columns (last 2 columns in V2 template)
        # Look in header row 2 for "AMOUNT" and "PERCENT" under "TOTAL"
        total_amount_col = None
        total_percent_col = None

        header_row = self.data_start_row - 1  # Row before data starts
        for col_idx in range(1, self.ws.max_column + 1):
            cell_value = self.ws.cell(row=header_row, column=col_idx).value
            if cell_value:
                cell_str = str(cell_value).upper().strip()
                if 'AMOUNT' in cell_str and total_amount_col is None:
                    # Check if this is under TOTAL section (not APPROVED CONTRACT)
                    if col_idx > 5:  # TOTAL section is after column E
                        total_amount_col = col_idx
                elif 'PERCENT' in cell_str:
                    total_percent_col = col_idx

        if not total_amount_col or not total_percent_col:
            logger.warning("Could not find TOTAL AMOUNT/PERCENT columns, using defaults")
            # Fallback to rightmost columns
            total_amount_col = self.ws.max_column - 1
            total_percent_col = self.ws.max_column

        logger.info(f"Using columns for totals - Amount: {total_amount_col}, Percent: {total_percent_col}")

        for row_idx in range(self.data_start_row, self.ws.max_row + 1):
            # Get cell values
            boq_code = self._get_cell_value('A', row_idx)
            description = self._get_cell_value('B', row_idx)

            # Skip empty rows
            if not boq_code:
                continue

            # Skip division header rows (DIV. 1, DIV. 2, etc.)
            if str(boq_code).upper().startswith('DIV.'):
                current_division = str(description).strip() if description else None
                continue

            # Check for TOTAL row - capture its amount and skip processing as BOQ item
            if str(boq_code).upper() == 'TOTAL' or (description and 'TOTAL' in str(description).upper() and 'SUB-TOTAL' not in str(description).upper()):
                # This is the main TOTAL row - extract the total amount
                try:
                    total_row_amount = self._get_cell_value_by_col(total_amount_col, row_idx)
                    if total_row_amount and total_row_amount != '' and total_row_amount != 0:
                        self.summary['total_period_amount_from_excel'] = self._parse_decimal(total_row_amount, row_idx, 'Total Amount')
                        logger.info(f"Found TOTAL row at row {row_idx} with amount: ₱{self.summary['total_period_amount_from_excel']:,.2f}")
                except Exception as e:
                    logger.warning(f"Could not extract total from TOTAL row: {str(e)}")
                continue

            # Skip subtotal rows
            if description and 'SUB-TOTAL' in str(description).upper():
                continue

            # Get period amount and percent from the TOTAL columns
            period_amount = self._get_cell_value_by_col(total_amount_col, row_idx)
            period_percent = self._get_cell_value_by_col(total_percent_col, row_idx)

            # Skip if no progress entered (both amount and percent are empty/zero)
            if (period_amount is None or period_amount == '' or period_amount == 0) and \
               (period_percent is None or period_percent == '' or period_percent == 0):
                continue

            try:
                # Parse period amount
                period_amount_value = self._parse_decimal(period_amount, row_idx, 'Period Amount')

                # Parse period percent - Excel stores percentages as decimals (0.019 = 1.9%)
                if isinstance(period_percent, (int, float)):
                    # It's a decimal from Excel formula, convert to percentage
                    period_percent_value = Decimal(str(period_percent * 100))
                else:
                    # Parse as normal
                    period_percent_value = self._parse_decimal(period_percent, row_idx, 'Period %')

                # Validate percent is reasonable
                if period_percent_value < 0 or period_percent_value > 100:
                    self.errors.append(f"Row {row_idx}: Period % must be between 0-100 (got {period_percent_value})")
                    continue

                # Get other values from standard columns
                quantity = self._get_cell_value('C', row_idx) or 0
                uom = self._get_cell_value('D', row_idx) or ''
                approved_amount = self._parse_decimal(self._get_cell_value('E', row_idx), row_idx, 'Approved Amount')

                # For V2 template, we don't have previous cumulative stored
                # We'll set cumulative = period (this is weekly progress)
                cumulative_percent = period_percent_value
                cumulative_amount = period_amount_value
                previous_percent = Decimal('0')
                previous_amount = Decimal('0')

                # Validate amount doesn't exceed approved
                if period_amount_value > approved_amount * Decimal('1.01'):  # Allow 1% tolerance
                    self.warnings.append(
                        f"Row {row_idx} ({boq_code}): Period amount (₱{period_amount_value:,.2f}) "
                        f"exceeds approved amount (₱{approved_amount:,.2f})"
                    )

                # Add BOQ item
                self.boq_items.append({
                    'boq_item_code': str(boq_code).strip(),
                    'description': str(description).strip() if description else '',
                    'division': current_division or '',
                    'task_group': '',  # V2 template doesn't have task column
                    'quantity': self._parse_decimal(quantity, row_idx, 'Quantity') if quantity else Decimal('0'),
                    'uom': str(uom).strip(),
                    'approved_amount': approved_amount,
                    'previous_cumulative_percent': previous_percent,
                    'cumulative_percent': cumulative_percent,
                    'cumulative_amount': cumulative_amount,
                    'period_progress_percent': period_percent_value,
                    'period_progress_amount': period_amount_value,
                    'progress_decreased': False,  # Not tracked in V2
                    'remarks': '',  # V2 doesn't have remarks column
                    'row_number': row_idx
                })

                self.summary['total_items'] += 1
                self.summary['items_with_progress'] += 1
                # Don't sum percentages here - use row 3 value in _calculate_summary instead
                self.summary['total_period_amount'] += period_amount_value

            except (ValueError, InvalidOperation) as e:
                self.errors.append(f"Row {row_idx}: Invalid data - {str(e)}")
                continue

    def _calculate_summary(self):
        """Calculate summary statistics"""
        if not self.boq_items:
            self.warnings.append("No BOQ items with progress found in the Excel file")
            return

        # Calculate total approved budget from all BOQ items
        total_approved = Decimal('0')
        for item in self.boq_items:
            total_approved += item.get('approved_amount', Decimal('0'))

        self.summary['total_approved_amount'] = total_approved

        # Use the TOTAL row amount from Excel if available (more accurate than summing individual items)
        if 'total_period_amount_from_excel' in self.summary and self.summary['total_period_amount_from_excel'] > 0:
            self.summary['total_period_amount'] = self.summary['total_period_amount_from_excel']
            logger.info(f"Using TOTAL row amount from Excel: ₱{self.summary['total_period_amount']:,.2f}")
        else:
            # Fallback: use the summed amount from individual BOQ items
            logger.warning(f"TOTAL row not found, using summed amount: ₱{self.summary['total_period_amount']:,.2f}")

        # Use the weekly progress from row 3 instead of summing BOQ item percentages
        # Row 3 contains the correct calculated percentage (PROGRESS THIS WEEK)
        if 'weekly_progress_percent' in self.summary:
            self.summary['total_period_percent'] = self.summary['weekly_progress_percent']
            logger.info(f"Using row 3 weekly progress: {self.summary['weekly_progress_percent']}%")
        else:
            # Fallback: calculate from total_period_amount / total_approved_amount
            if total_approved > 0 and self.summary['total_period_amount'] > 0:
                calculated_percent = (self.summary['total_period_amount'] / total_approved) * 100
                self.summary['total_period_percent'] = Decimal(str(calculated_percent))
                logger.warning(f"Row 3 weekly progress not found, calculated: {calculated_percent}%")

    def _validate_data(self):
        """Validate extracted data"""
        if not self.boq_items:
            self.errors.append("No valid BOQ progress entries found in Excel file")

        # Check for duplicate BOQ codes
        boq_codes = [item['boq_item_code'] for item in self.boq_items]
        duplicates = set([code for code in boq_codes if boq_codes.count(code) > 1])

        if duplicates:
            self.errors.append(f"Duplicate BOQ codes found: {', '.join(duplicates)}")

    def _get_cell_value(self, column, row):
        """Get cell value safely by column letter"""
        try:
            cell = self.ws[f'{column}{row}']
            value = cell.value

            # Handle formulas
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                # Formula cell - use calculated value
                return value

            # Clean string values
            if isinstance(value, str):
                value = value.strip()
                if value == '':
                    return None

            return value

        except Exception as e:
            logger.warning(f"Error reading cell {column}{row}: {str(e)}")
            return None

    def _get_cell_value_by_col(self, column_index, row):
        """Get cell value safely by column index"""
        try:
            cell = self.ws.cell(row=row, column=column_index)
            value = cell.value

            # Handle formulas
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                # Formula cell - use calculated value
                return value

            # Clean string values
            if isinstance(value, str):
                value = value.strip()
                if value == '':
                    return None

            return value

        except Exception as e:
            logger.warning(f"Error reading cell at column {column_index}, row {row}: {str(e)}")
            return None

    def _parse_decimal(self, value, row_idx, field_name):
        """Parse value to Decimal"""
        if value is None or value == '':
            return Decimal('0')

        try:
            # Remove currency symbols and commas
            if isinstance(value, str):
                value = value.replace('₱', '').replace(',', '').strip()

                # Handle percentage strings
                if '%' in value:
                    value = value.replace('%', '').strip()

            return Decimal(str(value))

        except (ValueError, InvalidOperation) as e:
            raise ValueError(f"{field_name} has invalid format: {value}")


def read_progress_excel(file_path_or_stream):
    """
    Read and validate progress Excel file

    Args:
        file_path_or_stream: File path or file-like object

    Returns:
        dict: Validation result with BOQ items and summary
    """
    reader = ProgressExcelReader(file_path_or_stream)
    return reader.read_and_validate()
