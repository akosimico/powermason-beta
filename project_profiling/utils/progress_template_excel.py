"""
Progress Template Excel Generator
Creates blank Excel templates for offline progress entry
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime


class ProgressTemplateExcelGenerator:
    """Generate blank Excel template for PM to fill offline"""

    def __init__(self, template_data):
        """
        Initialize with template data from progress_template_generator

        Args:
            template_data: Dict with project, divisions, tasks, BOQ items
        """
        self.template_data = template_data
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Progress Entry"

        # Styles
        self.header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        self.division_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.division_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        self.instruction_font = Font(name='Arial', size=10, italic=True, color='0066CC')
        self.editable_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        self.readonly_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.current_row = 1

    def generate_template(self):
        """
        Generate blank Excel template for progress entry

        Returns:
            Workbook: Excel workbook ready for download
        """
        self._add_header()
        self._add_instructions()
        self._add_summary_section()
        self._add_column_headers()
        self._add_boq_items()
        self._add_data_validations()
        self._format_columns()
        self._protect_cells()

        return self.wb

    def _add_header(self):
        """Add template header"""
        # Project name
        self.ws[f'A{self.current_row}'] = self.template_data['project_name']
        self.ws[f'A{self.current_row}'].font = Font(name='Arial', size=16, bold=True)
        self.current_row += 1

        # Template title
        self.ws[f'A{self.current_row}'] = "Weekly Progress Report - Entry Template"
        self.ws[f'A{self.current_row}'].font = Font(name='Arial', size=14, bold=True, color='0066CC')
        self.current_row += 1

        # Week info
        week_text = f"Week: {self.template_data['week_start_date']} to {self.template_data['week_end_date']}"
        self.ws[f'A{self.current_row}'] = week_text
        self.ws[f'A{self.current_row}'].font = Font(name='Arial', size=11, bold=True)
        self.current_row += 2

    def _add_instructions(self):
        """Add usage instructions"""
        instructions = [
            "INSTRUCTIONS:",
            "1. Fill in YELLOW cells only (Cumulative %, Cumulative Amount, Remarks)",
            "2. Enter CUMULATIVE progress (total completion), NOT incremental",
            "3. Cumulative % must be between 0-100",
            "4. If progress decreased, provide explanation in Remarks",
            "5. Leave Amount blank to auto-calculate (% × Approved Amount)",
            "6. Save this file when done",
            "7. Upload back to the system for submission",
            "",
            "GRAY cells are read-only and will be auto-filled by the system"
        ]

        for instruction in instructions:
            cell = self.ws[f'A{self.current_row}']
            cell.value = instruction
            cell.font = self.instruction_font
            self.current_row += 1

        self.current_row += 1

    def _add_summary_section(self):
        """Add summary section (will be auto-calculated on upload)"""
        self.ws[f'A{self.current_row}'] = "SUMMARY (Auto-calculated on upload)"
        self.ws[f'A{self.current_row}'].font = Font(name='Arial', size=12, bold=True)
        self.ws[f'A{self.current_row}'].fill = self.readonly_fill
        self.current_row += 1

        summary_labels = [
            'Total Period Progress:',
            'Total Period Amount:',
            'Total BOQ Items:',
            'Items with Progress:'
        ]

        for label in summary_labels:
            self.ws[f'A{self.current_row}'] = label
            self.ws[f'A{self.current_row}'].font = Font(name='Arial', size=10, bold=True)
            self.ws[f'B{self.current_row}'] = "Will be calculated"
            self.ws[f'B{self.current_row}'].fill = self.readonly_fill
            self.current_row += 1

        self.current_row += 1

    def _add_column_headers(self):
        """Add column headers"""
        headers = [
            ('BOQ Code', 'A'),
            ('Description', 'B'),
            ('Division', 'C'),
            ('Qty', 'D'),
            ('UOM', 'E'),
            ('Unit Price (₱)', 'F'),
            ('Approved Amount (₱)', 'G'),
            ('Previous %', 'H'),
            ('Cumulative % ✏', 'I'),  # Editable
            ('Cumulative Amount (₱) ✏', 'J'),  # Editable
            ('Remarks ✏', 'K'),  # Editable
        ]

        for header, col in headers:
            cell = self.ws[f'{col}{self.current_row}']
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        self.current_row += 1

    def _add_boq_items(self):
        """Add BOQ items from template data"""
        if not self.template_data.get('divisions'):
            # No items
            cell = self.ws[f'A{self.current_row}']
            cell.value = "No BOQ items found for this project"
            self.ws.merge_cells(f'A{self.current_row}:K{self.current_row}')
            return

        data_start_row = self.current_row

        for division in self.template_data['divisions']:
            # Division header
            cell = self.ws[f'A{self.current_row}']
            cell.value = division['name']
            cell.font = self.division_font
            cell.fill = self.division_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')

            # Merge division header
            self.ws.merge_cells(f'A{self.current_row}:K{self.current_row}')
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                self.ws[f'{col}{self.current_row}'].border = self.border

            self.current_row += 1

            # Add BOQ items directly from division
            for item in division['boq_items']:
                # BOQ Code
                cell = self.ws[f'A{self.current_row}']
                cell.value = item['code']
                cell.fill = self.readonly_fill
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left')

                # Description
                cell = self.ws[f'B{self.current_row}']
                cell.value = item['description']
                cell.fill = self.readonly_fill
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', wrap_text=True)

                # Division
                cell = self.ws[f'C{self.current_row}']
                cell.value = division['name']
                cell.fill = self.readonly_fill
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left')

                # Quantity
                cell = self.ws[f'D{self.current_row}']
                cell.value = float(item.get('quantity', 0))
                cell.fill = self.readonly_fill
                cell.border = self.border
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

                # UOM
                cell = self.ws[f'E{self.current_row}']
                cell.value = item.get('uom', '')
                cell.fill = self.readonly_fill
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')

                # Unit Price
                cell = self.ws[f'F{self.current_row}']
                cell.value = float(item.get('unit_price', 0))
                cell.fill = self.readonly_fill
                cell.border = self.border
                cell.number_format = '₱#,##0.00'
                cell.alignment = Alignment(horizontal='right')

                # Approved Amount
                cell = self.ws[f'G{self.current_row}']
                cell.value = float(item['approved_amount'])
                cell.fill = self.readonly_fill
                cell.border = self.border
                cell.number_format = '₱#,##0.00'
                cell.alignment = Alignment(horizontal='right')

                # Previous %
                cell = self.ws[f'H{self.current_row}']
                cell.value = float(item['previous_cumulative_percent'])
                cell.fill = self.readonly_fill
                cell.border = self.border
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal='right')

                # Cumulative % - EDITABLE
                cell = self.ws[f'I{self.current_row}']
                cell.value = float(item['previous_cumulative_percent'])  # Pre-fill with previous
                cell.fill = self.editable_fill
                cell.border = self.border
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')

                # Cumulative Amount - EDITABLE
                cell = self.ws[f'J{self.current_row}']
                cell.value = None  # Leave blank for auto-calculation
                cell.fill = self.editable_fill
                cell.border = self.border
                cell.number_format = '₱#,##0.00'
                cell.alignment = Alignment(horizontal='right')

                # Remarks - EDITABLE
                cell = self.ws[f'K{self.current_row}']
                cell.value = ""
                cell.fill = self.editable_fill
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left')

                self.current_row += 1

        # Store data range for later
        self.data_start_row = data_start_row
        self.data_end_row = self.current_row - 1

    def _add_data_validations(self):
        """Add data validation for editable cells"""
        if not hasattr(self, 'data_start_row'):
            return

        # Validation for Cumulative % (0-100)
        percent_validation = DataValidation(
            type="decimal",
            operator="between",
            formula1=0,
            formula2=100,
            allow_blank=False,
            showErrorMessage=True,
            error='Percentage must be between 0 and 100',
            errorTitle='Invalid Percentage'
        )
        percent_validation.prompt = 'Enter cumulative completion percentage (0-100)'
        percent_validation.promptTitle = 'Cumulative Progress'

        # Apply to column I (Cumulative %)
        percent_range = f'I{self.data_start_row}:I{self.data_end_row}'
        self.ws.add_data_validation(percent_validation)
        percent_validation.add(percent_range)

        # Validation for Amount (positive numbers)
        amount_validation = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1=0,
            allow_blank=True,
            showErrorMessage=True,
            error='Amount must be positive or leave blank for auto-calculation',
            errorTitle='Invalid Amount'
        )
        amount_validation.prompt = 'Enter cumulative amount or leave blank to auto-calculate'
        amount_validation.promptTitle = 'Cumulative Amount'

        # Apply to column J (Cumulative Amount)
        amount_range = f'J{self.data_start_row}:J{self.data_end_row}'
        self.ws.add_data_validation(amount_validation)
        amount_validation.add(amount_range)

    def _format_columns(self):
        """Set column widths"""
        column_widths = {
            'A': 12,  # BOQ Code
            'B': 45,  # Description
            'C': 25,  # Division
            'D': 10,  # Qty
            'E': 8,   # UOM
            'F': 15,  # Unit Price
            'G': 18,  # Approved Amount
            'H': 12,  # Previous %
            'I': 15,  # Cumulative % (editable)
            'J': 20,  # Cumulative Amount (editable)
            'K': 35,  # Remarks (editable)
        }

        for col, width in column_widths.items():
            self.ws.column_dimensions[col].width = width

    def _protect_cells(self):
        """Protect read-only cells, allow editing of yellow cells"""
        # Note: Full sheet protection would require password
        # For now, we just use visual cues (yellow vs gray)
        pass


def generate_blank_template(template_data):
    """
    Generate blank Excel template for PM to download

    Args:
        template_data: Dict from progress_template_generator

    Returns:
        Workbook: Excel workbook for download
    """
    generator = ProgressTemplateExcelGenerator(template_data)
    return generator.generate_template()
