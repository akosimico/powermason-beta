from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def create_hierarchical_boq_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value = "Bill of Quantities (Hierarchical Template)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    # Top total amount (auto-calculated from Amount column)
    ws["E2"].value = "Total Amount (PHP)"
    ws["E2"].font = Font(bold=True)
    # Sum only Division subtotals (level = 0)
    ws["F2"].value = "=SUMIF($G$10:$G$1000,0,$F$10:$F$1000)"
    ws["F2"].number_format = "#,##0.00"
    ws.merge_cells("G2:G2")

    # Project Info section
    project_info = [
        ("Project Name", "Residential House Construction - 2BR/2BA"),
        ("Lot Size (sqm)", "500"),
        ("Floor Area (sqm)", "120"),
        ("Project Location", "Metro Manila, Philippines"),
        ("Project Duration", "6 months"),
        ("Contractor", "ABC Construction Corp."),
    ]
    base_row = 4
    for idx, (label, val) in enumerate(project_info):
        r = base_row + idx
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)

    # Column headers
    headers = [
        "Code",
        "Description",
        "UOM",
        "Quantity",
        "Unit Cost",
        "Amount",
        "Level",
    ]
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Sample data demonstrating hierarchy and logic rules - Realistic Construction Project
    sample_rows = [
        ("DIV 1", "GENERAL REQUIREMENTS", "", "", "", ""),
        ("1.1", "Mobilization & Demobilization", "", "", "", ""),
        ("1.1.1", "Site mobilization and setup", "lot", 1, 75000, 75000),
        ("1.1.2", "Temporary facilities (office, storage, toilet)", "lot", 1, 120000, 120000),
        ("1.1.3", "Construction fence and security", "lm", 200, 450, 90000),
        ("1.1.4", "Site clearing and preparation", "sqm", 500, 85, 42500),
        ("1.2", "Project Management & Supervision", "", "", "", ""),
        ("1.2.1", "Project Manager (6 months)", "mo", 6, 45000, 270000),
        ("1.2.2", "Site Engineer (6 months)", "mo", 6, 35000, 210000),
        ("1.2.3", "Safety Officer (6 months)", "mo", 6, 28000, 168000),
        ("1.2.4", "Quality Control Inspector (6 months)", "mo", 6, 25000, 150000),
        ("1.3", "Permits & Licenses", "", "", "", ""),
        ("1.3.1", "Building permit and fees", "lot", 1, 45000, 45000),
        ("1.3.2", "Electrical permit", "lot", 1, 12000, 12000),
        ("1.3.3", "Plumbing permit", "lot", 1, 10000, 10000),
        ("1.3.4", "Environmental clearance", "lot", 1, 15000, 15000),

        ("DIV 2", "SITE PREPARATION & EXCAVATION", "", "", "", ""),
        ("2.1", "Earthworks", "", "", "", ""),
        ("2.1.1", "Excavation for foundation", "cum", 45, 850, 38250),
        ("2.1.2", "Backfilling and compaction", "cum", 35, 450, 15750),
        ("2.1.3", "Site grading and leveling", "sqm", 500, 120, 60000),
        ("2.2", "Foundation Preparation", "", "", "", ""),
        ("2.2.1", "Gravel bedding (150mm thick)", "cum", 8, 1200, 9600),
        ("2.2.2", "Vapor barrier (0.15mm PE)", "sqm", 120, 35, 4200),

        ("DIV 3", "CONCRETE WORKS", "", "", "", ""),
        ("3.1", "Foundation", "", "", "", ""),
        ("3.1.1", "Footing concrete (Class A)", "cum", 12, 4500, 54000),
        ("3.1.2", "Foundation wall concrete", "cum", 8, 4500, 36000),
        ("3.1.3", "Foundation reinforcement", "kg", 1200, 65, 78000),
        ("3.2", "Structural Elements", "", "", "", ""),
        ("3.2.1", "Ground floor slab (100mm)", "sqm", 120, 850, 102000),
        ("3.2.2", "Columns and beams", "cum", 15, 4500, 67500),
        ("3.2.3", "Structural reinforcement", "kg", 2500, 65, 162500),
        ("3.3", "Concrete Finishing", "", "", "", ""),
        ("3.3.1", "Concrete curing compound", "sqm", 200, 25, 5000),
        ("3.3.2", "Concrete sealer", "sqm", 200, 35, 7000),

        ("DIV 4", "MASONRY WORKS", "", "", "", ""),
        ("4.1", "Blockwork", "", "", "", ""),
        ("4.1.1", "CHB 6'' (150mm) walls", "sqm", 180, 950, 171000),
        ("4.1.2", "CHB 4'' (100mm) partition walls", "sqm", 120, 750, 90000),
        ("4.1.3", "Mortar mix (1:3 cement-sand)", "cum", 8, 2800, 22400),
        ("4.2", "Plastering", "", "", "", ""),
        ("4.2.1", "Cement plaster (20mm thick)", "sqm", 300, 180, 54000),
        ("4.2.2", "Smooth finish plaster", "sqm", 300, 45, 13500),
        ("4.3", "Tiling", "", "", "", ""),
        ("4.3.1", "Floor tiles (300x300mm)", "sqm", 80, 450, 36000),
        ("4.3.2", "Wall tiles (200x300mm)", "sqm", 60, 380, 22800),
        ("4.3.3", "Tile adhesive and grout", "sqm", 140, 85, 11900),

        ("DIV 5", "CARPENTRY & WOODWORKS", "", "", "", ""),
        ("5.1", "Structural Woodworks", "", "", "", ""),
        ("5.1.1", "Roof framing (2x6 lumber)", "lm", 200, 180, 36000),
        ("5.1.2", "Ceiling joists (2x4 lumber)", "lm", 150, 120, 18000),
        ("5.1.3", "Plywood sheathing (12mm)", "sqm", 100, 450, 45000),
        ("5.2", "Doors and Windows", "", "", "", ""),
        ("5.2.1", "Main door (solid wood)", "pc", 1, 8500, 8500),
        ("5.2.2", "Interior doors (hollow core)", "pc", 6, 2500, 15000),
        ("5.2.3", "Window frames (aluminum)", "pc", 8, 3200, 25600),
        ("5.2.4", "Glass panels (6mm clear)", "sqm", 25, 280, 7000),
        ("5.3", "Cabinetry", "", "", "", ""),
        ("5.3.1", "Kitchen cabinets (custom)", "lm", 8, 3500, 28000),
        ("5.3.2", "Bathroom vanity", "pc", 2, 4500, 9000),

        ("DIV 6", "ROOFING WORKS", "", "", "", ""),
        ("6.1", "Roof Structure", "", "", "", ""),
        ("6.1.1", "Roof trusses (prefabricated)", "set", 12, 2500, 30000),
        ("6.1.2", "Roof purlins (C-purlins)", "lm", 80, 180, 14400),
        ("6.2", "Roof Covering", "", "", "", ""),
        ("6.2.1", "Galvanized iron sheets (0.4mm)", "sqm", 120, 450, 54000),
        ("6.2.2", "Roof insulation (50mm)", "sqm", 120, 180, 21600),
        ("6.2.3", "Gutters and downspouts", "lm", 50, 280, 14000),
        ("6.3", "Roof Accessories", "", "", "", ""),
        ("6.3.1", "Ridge caps and flashings", "lm", 30, 180, 5400),
        ("6.3.2", "Roof vents", "pc", 4, 850, 3400),

        ("DIV 7", "ELECTRICAL WORKS", "", "", "", ""),
        ("7.1", "Electrical Rough-in", "", "", "", ""),
        ("7.1.1", "Main electrical panel (100A)", "pc", 1, 8500, 8500),
        ("7.1.2", "THHN wires (12 AWG)", "lm", 800, 45, 36000),
        ("7.1.3", "THHN wires (10 AWG)", "lm", 200, 65, 13000),
        ("7.1.4", "PVC conduits (20mm)", "lm", 600, 25, 15000),
        ("7.1.5", "PVC conduits (25mm)", "lm", 200, 35, 7000),
        ("7.1.6", "Utility boxes and covers", "pc", 45, 120, 5400),
        ("7.2", "Electrical Devices", "", "", "", ""),
        ("7.2.1", "Circuit breakers (20A)", "pc", 12, 450, 5400),
        ("7.2.2", "Convenience outlets (GFCI)", "pc", 25, 350, 8750),
        ("7.2.3", "Light switches (single pole)", "pc", 15, 180, 2700),
        ("7.2.4", "Light switches (3-way)", "pc", 8, 280, 2240),
        ("7.3", "Lighting Fixtures", "", "", "", ""),
        ("7.3.1", "LED downlights (12W)", "pc", 20, 650, 13000),
        ("7.3.2", "LED ceiling lights (18W)", "pc", 8, 850, 6800),
        ("7.3.3", "LED outdoor lights", "pc", 4, 1200, 4800),
        ("7.3.4", "Ceiling fans with lights", "pc", 6, 1800, 10800),

        ("DIV 8", "PLUMBING WORKS", "", "", "", ""),
        ("8.1", "Water Supply System", "", "", "", ""),
        ("8.1.1", "PPR pipes (20mm)", "lm", 150, 85, 12750),
        ("8.1.2", "PPR pipes (25mm)", "lm", 80, 120, 9600),
        ("8.1.3", "PPR fittings and valves", "set", 1, 3500, 3500),
        ("8.1.4", "Water meter and connections", "set", 1, 2500, 2500),
        ("8.2", "Sanitary & Drainage System", "", "", "", ""),
        ("8.2.1", "uPVC pipes (110mm)", "lm", 120, 95, 11400),
        ("8.2.2", "uPVC pipes (75mm)", "lm", 80, 65, 5200),
        ("8.2.3", "uPVC fittings and traps", "set", 1, 1800, 1800),
        ("8.2.4", "Floor drains and cleanouts", "pc", 8, 450, 3600),
        ("8.3", "Plumbing Fixtures", "", "", "", ""),
        ("8.3.1", "Water closet (one-piece)", "pc", 2, 3500, 7000),
        ("8.3.2", "Lavatory sink with faucet", "pc", 2, 2800, 5600),
        ("8.3.3", "Kitchen sink with faucet", "pc", 1, 4500, 4500),
        ("8.3.4", "Shower set with mixer", "pc", 1, 3200, 3200),
        ("8.3.5", "Water heater (50L)", "pc", 1, 8500, 8500),

        ("DIV 9", "PAINTING WORKS", "", "", "", ""),
        ("9.1", "Surface Preparation", "", "", "", ""),
        ("9.1.1", "Wall cleaning and sanding", "sqm", 300, 25, 7500),
        ("9.1.2", "Primer application", "sqm", 300, 35, 10500),
        ("9.2", "Paint Application", "", "", "", ""),
        ("9.2.1", "Interior paint (latex)", "sqm", 300, 85, 25500),
        ("9.2.2", "Exterior paint (weatherproof)", "sqm", 150, 120, 18000),
        ("9.2.3", "Paint supplies and tools", "lot", 1, 2500, 2500),

        ("DIV 10", "FINISHING WORKS", "", "", "", ""),
        ("10.1", "Flooring", "", "", "", ""),
        ("10.1.1", "Vinyl flooring (living areas)", "sqm", 60, 450, 27000),
        ("10.1.2", "Ceramic tiles (wet areas)", "sqm", 20, 380, 7600),
        ("10.1.3", "Floor leveling compound", "sqm", 80, 65, 5200),
        ("10.2", "Ceiling Works", "", "", "", ""),
        ("10.2.1", "Gypsum board ceiling", "sqm", 100, 280, 28000),
        ("10.2.2", "Ceiling cornices", "lm", 80, 120, 9600),
        ("10.3", "Hardware & Accessories", "", "", "", ""),
        ("10.3.1", "Door hardware (handles, locks)", "set", 7, 450, 3150),
        ("10.3.2", "Window hardware", "set", 8, 280, 2240),
        ("10.3.3", "Safety grills", "sqm", 25, 350, 8750),
    ]

    start_row = 10
    for r_offset, row in enumerate(sample_rows):
        row_idx = start_row + r_offset
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=c_idx, value=value)
            cell.border = border
        # Level formula based on dot count in Code (column A)
        ws.cell(row=row_idx, column=7, value=f"=LEN(A{row_idx})-LEN(SUBSTITUTE(A{row_idx},'.',''))")
        # Styling for top-level DIV rows: bold and uppercase code
        code_val = str(row[0])
        if code_val.upper().startswith("DIV ") and "." not in code_val:
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=2).font = Font(bold=True)
            # Color division row (A:F) pale yellow
            div_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).fill = div_fill
            # Inline subtotal on same division row: label in E, numeric in F
            ws.cell(row=row_idx, column=5, value="TOTAL:").font = Font(bold=True)
            # Children codes start with the numeric prefix (e.g., 1.) so strip 'DIV ' and add a dot
            ws.cell(row=row_idx, column=6, value=f"=SUMPRODUCT((LEFT($A$10:$A$1000,LEN(SUBSTITUTE(A{row_idx},\"DIV \",\"\"))+1)=SUBSTITUTE(A{row_idx},\"DIV \",\"\")&\".\")*($G$10:$G$1000=2)*($F$10:$F$1000))")
            ws.cell(row=row_idx, column=6).number_format = "#,##0.00"
        # Styling for single-dot TASK rows: bold description
        if code_val.count('.') == 1:
            ws.cell(row=row_idx, column=2).font = Font(bold=True)

    # Inline subtotals applied directly on division rows above; no separate subtotal rows

    # Widths
    widths = [18, 50, 12, 12, 14, 16, 8]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
    # Hide helper Level column
    ws.column_dimensions['G'].hidden = True

    # Notes section
    notes = (
        "Notes:"\
        "\n- Use 'DIV X' for Division (no dot) - Major work categories"\
        "\n- Use 'DIV X.Y' for Task (one dot) - Sub-categories within divisions"\
        "\n- Use 'DIV X.Y.Z' for Materials/Requirements (two dots) - Specific items with quantities"\
        "\n- Quantities are based on 120 sqm floor area residential house"\
        "\n- Unit costs are current market rates in PHP (Philippine Peso)"\
        "\n- All amounts are automatically calculated (Quantity × Unit Cost)"\
        "\n- Division totals are automatically calculated from their sub-items"\
        "\n- Project total excludes VAT and contingency (to be added separately)"
    )
    notes_row = start_row + len(sample_rows) + 2
    ws.merge_cells(f"A{notes_row}:G{notes_row + 4}")
    ncell = ws[f"A{notes_row}"]
    ncell.value = notes
    ncell.alignment = Alignment(wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def create_electrical_boq_template() -> bytes:
    """Create specialized BOQ template for Electrical Works"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Electrical BOQ"

    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value = "Electrical Works - Bill of Quantities"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    # Project Info
    project_info = [
        ("Project Name", "Residential House Electrical Installation"),
        ("Lot Size (sqm)", "500"),
        ("Floor Area (sqm)", "200"),
        ("Project Location", "Metro Manila, Philippines"),
        ("Contractor", "ABC Electrical Corp."),
    ]
    base_row = 4
    for idx, (label, val) in enumerate(project_info):
        r = base_row + idx
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)

    # Column headers
    headers = ["Code", "Description", "UOM", "Quantity", "Unit Cost", "Amount", "Level"]
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Electrical Works Sample Data
    sample_rows = [
        ("DIV 1", "GENERAL REQUIREMENTS", "", "", "", ""),
        ("1.1", "Mobilization & Demobilization", "", "", "", ""),
        ("1.1.1", "Site mobilization and setup", "lot", 1, 15000, 15000),
        ("1.1.2", "Temporary electrical connection", "lot", 1, 8000, 8000),
        ("1.1.3", "Construction lighting", "lot", 1, 5000, 5000),
        ("1.1.4", "Site demobilization", "lot", 1, 8000, 8000),
        ("1.2", "Project Management & Supervision", "", "", "", ""),
        ("1.2.1", "Person In Charge (3 months)", "mo", 3, 35000, 105000),
        ("1.2.2", "Quality Officer (3 months)", "mo", 3, 25000, 75000),
        ("1.2.3", "Foreman (1 month)", "mo", 1, 20000, 20000),
        ("1.3", "Permits & Licenses", "", "", "", ""),
        ("1.3.1", "Electrical permit", "lot", 1, 12000, 12000),
        ("1.3.2", "Electrical inspection fees", "lot", 1, 5000, 5000),
        ("1.3.3", "Electrical contractor license", "lot", 1, 3000, 3000),
        ("1.4", "Safety & Quality Control", "", "", "", ""),
        ("1.4.1", "Safety equipment and PPE", "lot", 1, 8000, 8000),
        ("1.4.2", "Quality control testing", "lot", 1, 10000, 10000),
        ("1.4.3", "Electrical testing instruments", "lot", 1, 15000, 15000),

        ("DIV 2", "POWER DISTRIBUTION SYSTEM", "", "", "", ""),
        ("2.1", "Main Electrical Panel", "", "", "", ""),
        ("2.1.1", "Main distribution board (100A)", "pc", 1, 25000, 25000),
        ("2.1.2", "Sub-distribution boards (63A)", "pc", 2, 12000, 24000),
        ("2.1.3", "Circuit breakers (MCB 20A)", "pc", 12, 850, 10200),
        ("2.1.4", "Circuit breakers (MCB 32A)", "pc", 6, 1200, 7200),
        ("2.1.5", "RCD/GFCI breakers (30mA)", "pc", 4, 2500, 10000),
        ("2.2", "Power Cables & Wiring", "", "", "", ""),
        ("2.2.1", "THHN cables (12 AWG)", "lm", 500, 45, 22500),
        ("2.2.2", "THHN cables (10 AWG)", "lm", 300, 65, 19500),
        ("2.2.3", "THHN cables (8 AWG)", "lm", 150, 95, 14250),
        ("2.2.4", "THHN cables (6 AWG)", "lm", 100, 150, 15000),
        ("2.2.5", "Grounding cables (6 AWG)", "lm", 200, 85, 17000),
        ("2.3", "Conduits & Raceways", "", "", "", ""),
        ("2.3.1", "PVC conduits (20mm)", "lm", 800, 25, 20000),
        ("2.3.2", "PVC conduits (25mm)", "lm", 400, 35, 14000),
        ("2.3.3", "PVC conduits (32mm)", "lm", 200, 50, 10000),
        ("2.3.4", "Steel conduits (25mm)", "lm", 100, 120, 12000),

        ("DIV 3", "LIGHTING SYSTEM", "", "", "", ""),
        ("3.1", "Indoor Lighting", "", "", "", ""),
        ("3.1.1", "LED ceiling lights (18W)", "pc", 15, 850, 12750),
        ("3.1.2", "LED downlights (12W)", "pc", 20, 650, 13000),
        ("3.1.3", "LED strip lights (18W)", "lm", 50, 180, 9000),
        ("3.1.4", "Emergency exit lights", "pc", 4, 1500, 6000),
        ("3.1.5", "Emergency lighting batteries", "pc", 4, 800, 3200),
        ("3.2", "Outdoor Lighting", "", "", "", ""),
        ("3.2.1", "LED floodlights (30W)", "pc", 4, 1800, 7200),
        ("3.2.2", "LED wall packs (20W)", "pc", 6, 1200, 7200),
        ("3.2.3", "LED bollard lights", "pc", 2, 2200, 4400),
        ("3.3", "Lighting Controls", "", "", "", ""),
        ("3.3.1", "Light switches (single pole)", "pc", 15, 180, 2700),
        ("3.3.2", "Light switches (3-way)", "pc", 8, 280, 2240),
        ("3.3.3", "Dimmer switches", "pc", 4, 450, 1800),
        ("3.3.4", "Motion sensors", "pc", 6, 350, 2100),

        ("DIV 4", "POWER OUTLETS & DEVICES", "", "", "", ""),
        ("4.1", "Power Outlets", "", "", "", ""),
        ("4.1.1", "GFCI outlets (20A)", "pc", 8, 350, 2800),
        ("4.1.2", "Standard outlets (15A)", "pc", 25, 180, 4500),
        ("4.1.3", "USB outlets", "pc", 6, 450, 2700),
        ("4.1.4", "Floor outlets", "pc", 2, 650, 1300),
        ("4.2", "Data & Communication", "", "", "", ""),
        ("4.2.1", "Data outlets (Cat6)", "pc", 8, 280, 2240),
        ("4.2.2", "Telephone outlets", "pc", 4, 150, 600),
        ("4.2.3", "TV outlets", "pc", 6, 200, 1200),
        ("4.3", "Specialized Outlets", "", "", "", ""),
        ("4.3.1", "Range outlets (50A)", "pc", 1, 1200, 1200),
        ("4.3.2", "Dryer outlets (30A)", "pc", 1, 800, 800),
        ("4.3.3", "EV charging outlets (40A)", "pc", 1, 3500, 3500),

        ("DIV 5", "FIRE ALARM & SAFETY", "", "", "", ""),
        ("5.1", "Fire Alarm System", "", "", "", ""),
        ("5.1.1", "Smoke detectors (battery)", "pc", 8, 450, 3600),
        ("5.1.2", "Heat detectors", "pc", 4, 350, 1400),
        ("5.1.3", "Manual pull stations", "pc", 2, 280, 560),
        ("5.1.4", "Horn strobes", "pc", 4, 650, 2600),
        ("5.1.5", "Fire alarm cables", "lm", 200, 35, 7000),
        ("5.2", "Emergency Systems", "", "", "", ""),
        ("5.2.1", "Emergency generator (5kVA)", "pc", 1, 25000, 25000),
        ("5.2.2", "ATS (Automatic Transfer Switch)", "pc", 1, 8000, 8000),
        ("5.2.3", "Emergency lighting system", "lot", 1, 8000, 8000),

        ("DIV 6", "SPECIALIZED SYSTEMS", "", "", "", ""),
        ("6.1", "HVAC Electrical", "", "", "", ""),
        ("6.1.1", "Aircon electrical connections", "pc", 4, 2500, 10000),
        ("6.1.2", "Exhaust fan connections", "pc", 6, 800, 4800),
        ("6.1.3", "Water heater connections", "pc", 2, 1200, 2400),
        ("6.2", "Security System", "", "", "", ""),
        ("6.2.1", "CCTV cameras", "pc", 4, 2500, 10000),
        ("6.2.2", "DVR system", "pc", 1, 8000, 8000),
        ("6.2.3", "Doorbell system", "pc", 1, 1500, 1500),
        ("6.3", "Home Automation", "", "", "", ""),
        ("6.3.1", "Smart switches", "pc", 8, 1200, 9600),
        ("6.3.2", "Smart outlets", "pc", 6, 800, 4800),
        ("6.3.3", "Home automation cables", "lm", 100, 45, 4500),
    ]

    start_row = 10
    for r_offset, row in enumerate(sample_rows):
        row_idx = start_row + r_offset
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=c_idx, value=value)
            cell.border = border
        # Level formula
        ws.cell(row=row_idx, column=7, value=f"=LEN(A{row_idx})-LEN(SUBSTITUTE(A{row_idx},'.',''))")
        # Styling for division rows
        code_val = str(row[0])
        if code_val.upper().startswith("DIV ") and "." not in code_val:
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=2).font = Font(bold=True)
            div_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).fill = div_fill
            ws.cell(row=row_idx, column=5, value="TOTAL:").font = Font(bold=True)
            ws.cell(row=row_idx, column=6, value=f"=SUMPRODUCT((LEFT($A$10:$A$1000,LEN(SUBSTITUTE(A{row_idx},\"DIV \",\"\"))+1)=SUBSTITUTE(A{row_idx},\"DIV \",\"\")&\".\")*($G$10:$G$1000=2)*($F$10:$F$1000))")
            ws.cell(row=row_idx, column=6).number_format = "#,##0.00"
        if code_val.count('.') == 1:
            ws.cell(row=row_idx, column=2).font = Font(bold=True)

    # Widths and formatting
    widths = [18, 50, 12, 12, 14, 16, 8]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
    ws.column_dimensions['G'].hidden = True

    # Notes
    notes = (
        "Notes:"\
        "\n- Specialized template for Electrical Works"\
        "\n- Includes power distribution, lighting, outlets, and safety systems"\
        "\n- Quantities based on 200 sqm residential house (500 sqm lot)"\
        "\n- Unit costs in PHP (Philippine Peso)"\
        "\n- All electrical work must comply with PEC (Philippine Electrical Code)"\
        "\n- Requires licensed electrical contractor for installation"\
        "\n- General Requirements include mobilization, permits, and supervision"
    )
    notes_row = start_row + len(sample_rows) + 2
    ws.merge_cells(f"A{notes_row}:G{notes_row + 4}")
    ncell = ws[f"A{notes_row}"]
    ncell.value = notes
    ncell.alignment = Alignment(wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def create_mechanical_boq_template() -> bytes:
    """Create specialized BOQ template for Mechanical Works"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mechanical BOQ"

    header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value = "Mechanical Works - Bill of Quantities"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    # Project Info
    project_info = [
        ("Project Name", "Residential House Mechanical Systems"),
        ("Lot Size (sqm)", "500"),
        ("Floor Area (sqm)", "200"),
        ("Project Location", "Metro Manila, Philippines"),
        ("Contractor", "ABC Mechanical Corp."),
    ]
    base_row = 4
    for idx, (label, val) in enumerate(project_info):
        r = base_row + idx
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)

    # Column headers
    headers = ["Code", "Description", "UOM", "Quantity", "Unit Cost", "Amount", "Level"]
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Mechanical Works Sample Data
    sample_rows = [
        ("DIV 1", "GENERAL REQUIREMENTS", "", "", "", ""),
        ("1.1", "Mobilization & Demobilization", "", "", "", ""),
        ("1.1.1", "Site mobilization and setup", "lot", 1, 12000, 12000),
        ("1.1.2", "Temporary water connection", "lot", 1, 5000, 5000),
        ("1.1.3", "Construction equipment rental", "lot", 1, 8000, 8000),
        ("1.1.4", "Site demobilization", "lot", 1, 6000, 6000),
        ("1.2", "Project Management & Supervision", "", "", "", ""),
        ("1.2.1", "Project Manager (3 months)", "mo", 3, 35000, 105000),
        ("1.2.2", "Foreman (3 months)", "mo", 3, 25000, 75000),
        ("1.2.3", "Person In Charge (1 month)", "mo", 1, 20000, 20000),
        ("1.3", "Permits & Licenses", "", "", "", ""),
        ("1.3.1", "Plumbing permit", "lot", 1, 10000, 10000),
        ("1.3.2", "Mechanical permit", "lot", 1, 8000, 8000),
        ("1.3.3", "Inspection fees", "lot", 1, 5000, 5000),
        ("1.4", "Safety & Quality Control", "", "", "", ""),
        ("1.4.1", "Safety equipment and PPE", "lot", 1, 6000, 6000),
        ("1.4.2", "Quality control testing", "lot", 1, 8000, 8000),
        ("1.4.3", "Testing instruments", "lot", 1, 12000, 12000),

        ("DIV 2", "HVAC SYSTEM", "", "", "", ""),
        ("2.1", "Air Conditioning Units", "", "", "", ""),
        ("2.1.1", "Split type aircon (1.5HP)", "pc", 4, 18000, 72000),
        ("2.1.2", "Split type aircon (2.0HP)", "pc", 2, 25000, 50000),
        ("2.1.3", "Window type aircon (1.0HP)", "pc", 2, 12000, 24000),
        ("2.1.4", "Aircon installation", "pc", 8, 3000, 24000),
        ("2.2", "Ventilation System", "", "", "", ""),
        ("2.2.1", "Exhaust fans (ceiling mounted)", "pc", 4, 1200, 4800),
        ("2.2.2", "Exhaust fans (wall mounted)", "pc", 6, 800, 4800),
        ("2.2.3", "Kitchen exhaust hood", "pc", 1, 8000, 8000),
        ("2.2.4", "Bathroom exhaust fans", "pc", 4, 600, 2400),
        ("2.3", "Air Distribution", "", "", "", ""),
        ("2.3.1", "Aircon electrical connections", "pc", 8, 1500, 12000),
        ("2.3.2", "Aircon drainage", "pc", 8, 800, 6400),
        ("2.3.3", "Aircon mounting brackets", "pc", 8, 500, 4000),

        ("DIV 3", "PLUMBING SYSTEM", "", "", "", ""),
        ("3.1", "Water Supply", "", "", "", ""),
        ("3.1.1", "PPR pipes (20mm)", "lm", 150, 85, 12750),
        ("3.1.2", "PPR pipes (25mm)", "lm", 100, 120, 12000),
        ("3.1.3", "PPR pipes (32mm)", "lm", 50, 150, 7500),
        ("3.1.4", "PPR fittings and valves", "set", 1, 4000, 4000),
        ("3.1.5", "Water meter (25mm)", "pc", 1, 3500, 3500),
        ("3.1.6", "Water pressure tank", "pc", 1, 8000, 8000),
        ("3.2", "Drainage System", "", "", "", ""),
        ("3.2.1", "uPVC pipes (110mm)", "lm", 100, 95, 9500),
        ("3.2.2", "uPVC pipes (75mm)", "lm", 80, 65, 5200),
        ("3.2.3", "uPVC pipes (50mm)", "lm", 60, 45, 2700),
        ("3.2.4", "uPVC fittings and traps", "set", 1, 2500, 2500),
        ("3.2.5", "Floor drains", "pc", 6, 450, 2700),
        ("3.3", "Plumbing Fixtures", "", "", "", ""),
        ("3.3.1", "Water closets (floor mounted)", "pc", 3, 3500, 10500),
        ("3.3.2", "Lavatory sinks", "pc", 4, 2500, 10000),
        ("3.3.3", "Kitchen sinks", "pc", 1, 4500, 4500),
        ("3.3.4", "Shower sets", "pc", 2, 3500, 7000),
        ("3.3.5", "Water heaters (50L)", "pc", 2, 8000, 16000),

        ("DIV 4", "FIRE PROTECTION SYSTEM", "", "", "", ""),
        ("4.1", "Fire Safety Equipment", "", "", "", ""),
        ("4.1.1", "Fire extinguishers (5kg)", "pc", 4, 800, 3200),
        ("4.1.2", "Fire extinguisher cabinets", "pc", 4, 500, 2000),
        ("4.1.3", "Fire blankets", "pc", 2, 300, 600),
        ("4.1.4", "Smoke detectors", "pc", 6, 400, 2400),
        ("4.2", "Emergency Equipment", "", "", "", ""),
        ("4.2.1", "Emergency lighting", "pc", 4, 1200, 4800),
        ("4.2.2", "Fire exit signs", "pc", 4, 600, 2400),
        ("4.2.3", "First aid kits", "pc", 2, 800, 1600),

        ("DIV 5", "WATER TREATMENT & STORAGE", "", "", "", ""),
        ("5.1", "Water Storage", "", "", "", ""),
        ("5.1.1", "Water storage tank (1000L)", "pc", 1, 8000, 8000),
        ("5.1.2", "Water pump (1HP)", "pc", 1, 12000, 12000),
        ("5.1.3", "Water pressure switch", "pc", 1, 1500, 1500),
        ("5.1.4", "Water level controller", "pc", 1, 2000, 2000),
        ("5.2", "Water Treatment", "", "", "", ""),
        ("5.2.1", "Water filter system", "pc", 1, 5000, 5000),
        ("5.2.2", "Water softener", "pc", 1, 8000, 8000),
        ("5.2.3", "UV sterilizer", "pc", 1, 6000, 6000),

        ("DIV 6", "MAINTENANCE & TESTING", "", "", "", ""),
        ("6.1", "System Testing", "", "", "", ""),
        ("6.1.1", "Plumbing pressure testing", "lot", 1, 3000, 3000),
        ("6.1.2", "HVAC performance testing", "lot", 1, 4000, 4000),
        ("6.1.3", "System commissioning", "lot", 1, 5000, 5000),
        ("6.2", "Maintenance", "", "", "", ""),
        ("6.2.1", "1-year maintenance contract", "lot", 1, 15000, 15000),
        ("6.2.2", "Spare parts kit", "lot", 1, 5000, 5000),
    ]

    start_row = 10
    for r_offset, row in enumerate(sample_rows):
        row_idx = start_row + r_offset
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=c_idx, value=value)
            cell.border = border
        # Level formula
        ws.cell(row=row_idx, column=7, value=f"=LEN(A{row_idx})-LEN(SUBSTITUTE(A{row_idx},'.',''))")
        # Styling for division rows
        code_val = str(row[0])
        if code_val.upper().startswith("DIV ") and "." not in code_val:
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=2).font = Font(bold=True)
            div_fill = PatternFill(start_color="E5F5E0", end_color="E5F5E0", fill_type="solid")
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).fill = div_fill
            ws.cell(row=row_idx, column=5, value="TOTAL:").font = Font(bold=True)
            ws.cell(row=row_idx, column=6, value=f"=SUMPRODUCT((LEFT($A$10:$A$1000,LEN(SUBSTITUTE(A{row_idx},\"DIV \",\"\"))+1)=SUBSTITUTE(A{row_idx},\"DIV \",\"\")&\".\")*($G$10:$G$1000=2)*($F$10:$F$1000))")
            ws.cell(row=row_idx, column=6).number_format = "#,##0.00"
        if code_val.count('.') == 1:
            ws.cell(row=row_idx, column=2).font = Font(bold=True)

    # Widths and formatting
    widths = [18, 50, 12, 12, 14, 16, 8]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
    ws.column_dimensions['G'].hidden = True

    # Notes
    notes = (
        "Notes:"\
        "\n- Specialized template for Mechanical Works"\
        "\n- Includes HVAC, plumbing, fire protection, and water systems"\
        "\n- Quantities based on 200 sqm residential house (500 sqm lot)"\
        "\n- Unit costs in PHP (Philippine Peso)"\
        "\n- All mechanical work must comply with local building codes"\
        "\n- Requires licensed mechanical engineer for design and supervision"\
        "\n- General Requirements include mobilization, permits, and supervision"
    )
    notes_row = start_row + len(sample_rows) + 2
    ws.merge_cells(f"A{notes_row}:G{notes_row + 4}")
    ncell = ws[f"A{notes_row}"]
    ncell.value = notes
    ncell.alignment = Alignment(wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def create_civil_boq_template() -> bytes:
    """Create specialized BOQ template for Civil Works"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Civil BOQ"

    header_fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")  # Blue
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value = "Civil Works - Bill of Quantities"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    # Project Info
    project_info = [
        ("Project Name", "Commercial Building Civil Works"),
        ("Lot Size (sqm)", "5000"),
        ("Floor Area (sqm)", "2000"),
        ("Project Location", "Metro Manila, Philippines"),
        ("Contractor", "ABC Civil Corp."),
    ]
    base_row = 4
    for idx, (label, val) in enumerate(project_info):
        r = base_row + idx
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)

    # Column headers
    headers = ["Code", "Description", "UOM", "Quantity", "Unit Cost", "Amount", "Level"]
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Civil Works Sample Data (now with General Requirements)
    sample_rows = [
        ("DIV 1", "GENERAL REQUIREMENTS", "", "", "", ""),
        ("1.1", "Mobilization & Temporary Facilities", "", "", "", ""),
        ("1.1.1", "Mobilization and setup", "lot", 1, 20000, 20000),
        ("1.1.2", "Temporary access road", "sqm", 200, 180, 36000),
        ("1.1.3", "Temporary fence and gate", "lm", 100, 450, 45000),
        ("1.1.4", "Temporary water & power connection", "lot", 1, 20000, 20000),
        ("1.2", "Project Supervision & Safety", "", "", "", ""),
        ("1.2.1", "Project Manager (3 months)", "mo", 3, 40000, 120000),
        ("1.2.2", "Safety officer (3 months)", "mo", 3, 25000, 75000),
        ("1.2.3", "Quality Assurance Officer", "lot", 1, 20000, 20000),
        ("1.3", "Permits & Documentation", "", "", "", ""),
        ("1.3.1", "Building permit", "lot", 1, 15000, 15000),
        ("1.3.2", "Safety and inspection fees", "lot", 1, 8000, 8000),
        ("1.3.3", "Environmental compliance", "lot", 1, 5000, 5000),

        ("DIV 2", "SITE PREPARATION", "", "", "", ""),
        ("2.1", "Earthworks", "", "", "", ""),
        ("2.1.1", "Site clearing and grubbing", "sqm", 2500, 45, 112500),
        ("2.1.2", "Excavation for foundation", "cum", 200, 850, 170000),
        ("2.1.3", "Backfilling and compaction", "cum", 150, 450, 67500),
        ("2.1.4", "Site grading and leveling", "sqm", 2500, 120, 300000),
        ("2.1.5", "Temporary drainage", "lm", 100, 250, 25000),

        ("DIV 3", "FOUNDATION WORKS", "", "", "", ""),
        ("3.1", "Concrete Works", "", "", "", ""),
        ("3.1.1", "Footing concrete (Class A)", "cum", 25, 4500, 112500),
        ("3.1.2", "Foundation wall concrete", "cum", 15, 4500, 67500),
        ("3.1.3", "Reinforcement bars", "kg", 3000, 65, 195000),
        ("3.1.4", "Formworks", "sqm", 200, 350, 70000),
        ("3.1.5", "Waterproofing works", "sqm", 300, 180, 54000),

        ("DIV 4", "STRUCTURAL WORKS", "", "", "", ""),
        ("4.1", "Reinforced Concrete", "", "", "", ""),
        ("4.1.1", "Ground floor slab (150mm)", "sqm", 2000, 850, 1700000),
        ("4.1.2", "Columns (300x300mm)", "cum", 20, 4500, 90000),
        ("4.1.3", "Beams (300x400mm)", "cum", 15, 4500, 67500),
        ("4.1.4", "Structural reinforcement", "kg", 5000, 65, 325000),
        ("4.1.5", "Formworks", "sqm", 500, 350, 175000),

        ("DIV 5", "MASONRY & WALLS", "", "", "", ""),
        ("5.1", "Blockwork", "", "", "", ""),
        ("5.1.1", "CHB 6'' (150mm) walls", "sqm", 400, 950, 380000),
        ("5.1.2", "CHB 4'' (100mm) partition walls", "sqm", 300, 750, 225000),
        ("5.1.3", "Plastering (20mm thick)", "sqm", 700, 180, 126000),
        ("5.1.4", "Wall paint (2 coats)", "sqm", 700, 85, 59500),
    ]

    start_row = 10
    for r_offset, row in enumerate(sample_rows):
        row_idx = start_row + r_offset
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=c_idx, value=value)
            cell.border = border
        # Level formula
        ws.cell(row=row_idx, column=7, value=f"=LEN(A{row_idx})-LEN(SUBSTITUTE(A{row_idx},'.',''))")
        # Styling for division rows
        code_val = str(row[0])
        if code_val.upper().startswith("DIV ") and "." not in code_val:
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=2).font = Font(bold=True)
            div_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).fill = div_fill
            ws.cell(row=row_idx, column=5, value="TOTAL:").font = Font(bold=True)
            ws.cell(row=row_idx, column=6, value=f"=SUMPRODUCT((LEFT($A$10:$A$1000,LEN(SUBSTITUTE(A{row_idx},\"DIV \",\"\"))+1)=SUBSTITUTE(A{row_idx},\"DIV \",\"\")&\".\")*($G$10:$G$1000=2)*($F$10:$F$1000))")
            ws.cell(row=row_idx, column=6).number_format = "#,##0.00"
        if code_val.count('.') == 1:
            ws.cell(row=row_idx, column=2).font = Font(bold=True)

    # Widths and formatting
    widths = [18, 50, 12, 12, 14, 16, 8]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
    ws.column_dimensions['G'].hidden = True

    # Notes
    notes = (
        "Notes:"\
        "\n- Specialized template for Civil Works"\
        "\n- Includes general requirements, site prep, foundation, structural, and masonry works"\
        "\n- Quantities based on 2000 sqm commercial building (5000 sqm lot)"\
        "\n- Unit costs in PHP (Philippine Peso)"\
        "\n- All civil works must comply with the National Building Code"\
        "\n- Requires licensed civil engineer for design and supervision"\
        "\n- General Requirements include mobilization, permits, and supervision"
    )
    notes_row = start_row + len(sample_rows) + 2
    ws.merge_cells(f"A{notes_row}:G{notes_row + 4}")
    ncell = ws[f"A{notes_row}"]
    ncell.value = notes
    ncell.alignment = Alignment(wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def create_architectural_boq_template() -> bytes:
    """Create specialized BOQ template for Architectural Works (with General Requirements)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Architectural BOQ"

    header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value = "Architectural Works - Bill of Quantities"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    # Project Info
    project_info = [
        ("Project Name", "Commercial Building Architectural Works"),
        ("Lot Size (sqm)", "3000"),
        ("Floor Area (sqm)", "2000"),
        ("Project Location", "Metro Manila, Philippines"),
        ("Contractor", "ABC Architectural Corp."),
    ]
    base_row = 4
    for idx, (label, val) in enumerate(project_info):
        r = base_row + idx
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)

    # Column headers
    headers = ["Code", "Description", "UOM", "Quantity", "Unit Cost", "Amount", "Level"]
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Architectural Works Sample Data (includes General Requirements)
    sample_rows = [
        # DIV 1 - General Requirements
        ("DIV 1", "GENERAL REQUIREMENTS", "", "", "", ""),
        ("1.1", "Mobilization & Demobilization", "", "", "", ""),
        ("1.1.1", "Site mobilization and setup", "lot", 1, 15000, 15000),
        ("1.1.2", "Temporary facilities and access", "lot", 1, 12000, 12000),
        ("1.1.3", "Site demobilization and cleanup", "lot", 1, 10000, 10000),
        ("1.2", "Project Management & Supervision", "", "", "", ""),
        ("1.2.1", "Quality Officer(3 months)", "mo", 3, 40000, 120000),
        ("1.2.2", "Site engineer / foreman (3 months)", "mo", 3, 25000, 75000),
        ("1.2.3", "Safety officer (2 months)", "mo", 2, 20000, 40000),
        ("1.3", "Permits & Licenses", "", "", "", ""),
        ("1.3.1", "Building permit", "lot", 1, 20000, 20000),
        ("1.3.2", "Occupancy permit", "lot", 1, 8000, 8000),
        ("1.4", "Safety & Quality Control", "", "", "", ""),
        ("1.4.1", "Safety equipment and PPE", "lot", 1, 10000, 10000),
        ("1.4.2", "Quality assurance testing", "lot", 1, 12000, 12000),

        # DIV 2 onwards (Architectural Works)
        ("DIV 2", "DOORS & WINDOWS", "", "", "", ""),
        ("2.1", "Doors", "", "", "", ""),
        ("2.1.1", "Main entrance door (aluminum)", "pc", 2, 15000, 30000),
        ("2.1.2", "Office doors (solid wood)", "pc", 20, 4500, 90000),
        ("2.1.3", "Conference room doors (glass)", "pc", 4, 8500, 34000),
        ("2.1.4", "Bathroom doors (hollow core)", "pc", 8, 2500, 20000),
        ("2.1.5", "Fire exit doors", "pc", 6, 3500, 21000),
        ("2.1.6", "Door hardware (handles, locks)", "set", 40, 450, 18000),
        ("2.2", "Windows", "", "", "", ""),
        ("2.2.1", "Aluminum windows (sliding)", "sqm", 80, 1200, 96000),
        ("2.2.2", "Aluminum windows (casement)", "sqm", 60, 1400, 84000),
        ("2.2.3", "Glass panels (6mm clear)", "sqm", 140, 280, 39200),
        ("2.2.4", "Window hardware", "set", 30, 280, 8400),
        ("2.2.5", "Window screens", "sqm", 140, 120, 16800),

        ("DIV 3", "CEILING & PARTITIONS", "", "", "", ""),
        ("3.1", "Ceiling Systems", "", "", "", ""),
        ("3.1.1", "Gypsum board ceiling", "sqm", 2000, 280, 560000),
        ("3.1.2", "Acoustic ceiling tiles", "sqm", 500, 180, 90000),
        ("3.1.3", "Ceiling cornices", "lm", 200, 120, 24000),
        ("3.1.4", "Ceiling lighting fixtures", "pc", 100, 350, 35000),
        ("3.1.5", "Ceiling access panels", "pc", 20, 450, 9000),
        ("3.2", "Partition Systems", "", "", "", ""),
        ("3.2.1", "Drywall partitions (100mm)", "sqm", 300, 350, 105000),
        ("3.2.2", "Glass partitions (frameless)", "sqm", 100, 1200, 120000),
        ("3.2.3", "Movable partitions", "sqm", 50, 850, 42500),
        ("3.2.4", "Partition doors", "pc", 15, 2800, 42000),

        ("DIV 4", "FLOORING & STAIRS", "", "", "", ""),
        ("4.1", "Floor Finishing", "", "", "", ""),
        ("4.1.1", "Ceramic tiles (300x300mm)", "sqm", 800, 450, 360000),
        ("4.1.2", "Marble tiles (600x600mm)", "sqm", 200, 850, 170000),
        ("4.1.3", "Wood flooring (engineered)", "sqm", 300, 650, 195000),
        ("4.2", "Stairs", "", "", "", ""),
        ("4.2.1", "Concrete stairs", "lm", 50, 1200, 60000),
        ("4.2.2", "Stair railings (stainless steel)", "lm", 50, 450, 22500),

        ("DIV 5", "FURNITURE & FIXTURES", "", "", "", ""),
        ("5.1", "Built-in Furniture", "", "", "", ""),
        ("5.1.1", "Reception desk", "pc", 1, 25000, 25000),
        ("5.1.2", "Office cabinets", "lm", 50, 1200, 60000),
        ("5.1.3", "Bathroom vanities", "pc", 8, 4500, 36000),

        ("DIV 6", "EXTERIOR ELEMENTS", "", "", "", ""),
        ("6.1", "Facade", "", "", "", ""),
        ("6.1.1", "Aluminum cladding", "sqm", 200, 850, 170000),
        ("6.1.2", "Glass curtain wall", "sqm", 300, 1200, 360000),
        ("6.1.3", "Facade lighting", "pc", 30, 1200, 36000),
    ]

    start_row = 10
    for r_offset, row in enumerate(sample_rows):
        row_idx = start_row + r_offset
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=c_idx, value=value)
            cell.border = border
        ws.cell(row=row_idx, column=7, value=f"=LEN(A{row_idx})-LEN(SUBSTITUTE(A{row_idx},'.',''))")

        code_val = str(row[0])
        # Highlight divisions
        if code_val.upper().startswith("DIV ") and "." not in code_val:
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=2).font = Font(bold=True)
            div_fill = PatternFill(start_color="F5E6D3", end_color="F5E6D3", fill_type="solid")
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).fill = div_fill
            ws.cell(row=row_idx, column=5, value="TOTAL:").font = Font(bold=True)
            ws.cell(row=row_idx, column=6, value=f"=SUMPRODUCT((LEFT($A$10:$A$1000,LEN(SUBSTITUTE(A{row_idx},\"DIV \",\"\"))+1)=SUBSTITUTE(A{row_idx},\"DIV \",\"\")&\".\")*($G$10:$G$1000=2)*($F$10:$F$1000))")
            ws.cell(row=row_idx, column=6).number_format = "#,##0.00"
        if code_val.count('.') == 1:
            ws.cell(row=row_idx, column=2).font = Font(bold=True)

    # Adjust widths
    widths = [18, 50, 12, 12, 14, 16, 8]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
    ws.column_dimensions['G'].hidden = True

    # Notes
    notes = (
        "Notes:"\
        "\n- Standard BOQ template for Architectural Works"\
        "\n- Includes general requirements, doors, windows, ceilings, flooring, and finishes"\
        "\n- Quantities based on 2000 sqm commercial building"\
        "\n- Unit costs in PHP (Philippine Peso)"\
        "\n- All works must comply with NBCP and local architectural standards"\
        "\n- Supervision must be done by a licensed architect"
    )
    notes_row = start_row + len(sample_rows) + 2
    ws.merge_cells(f"A{notes_row}:G{notes_row + 4}")
    ncell = ws[f"A{notes_row}"]
    ncell.value = notes
    ncell.alignment = Alignment(wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()