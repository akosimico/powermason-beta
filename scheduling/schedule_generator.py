"""
Excel Schedule Template Generator
Generates formatted Excel templates for project schedule planning
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.conf import settings


class ScheduleTemplateGenerator:
    """Generate Excel templates for project scheduling"""

    # Color scheme
    HEADER_COLOR = "1F4E78"  # Dark blue
    SCOPE_COLOR = "4472C4"   # Medium blue
    PROJECT_INFO_COLOR = "D9E1F2"  # Light blue

    def __init__(self, project):
        """
        Initialize generator with project data

        Args:
            project: ProjectProfile instance
        """
        self.project = project
        self.scopes = project.scopes.filter(is_deleted=False).order_by('id')
        self.wb = Workbook()

    def generate(self):
        """
        Generate complete Excel template

        Returns:
            str: Path to generated file
        """
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Create sheets
        self._create_instructions_sheet()
        self._create_schedule_sheet()

        # Save file
        filename = self._get_filename()
        filepath = os.path.join(settings.MEDIA_ROOT, 'schedule_templates', filename)

        # Ensure directory exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        self.wb.save(filepath)
        return filepath

    def _create_instructions_sheet(self):
        """Create instructions sheet with guide for PM"""
        ws = self.wb.create_sheet("Instructions", 0)

        instructions = [
            ("PROJECT SCHEDULE TEMPLATE - INSTRUCTIONS", True),
            ("", False),
            ("HOW TO USE THIS TEMPLATE:", True),
            ("", False),
            ("1. DO NOT modify the project information or scope names", False),
            ("2. Fill in task details under each SCOPE section:", False),
            ("   - ITEM: Sequential number (1.0, 2.0, 3.0, etc.)", False),
            ("   - ACTIVITY: Task name/description", False),
            ("   - START: Start date (format: DD-MMM-YY, e.g., 02-Aug-24)", False),
            ("   - END: End date (format: DD-MMM-YY)", False),
            ("   - DAYS: Duration in days (auto-calculated if left blank)", False),
            ("   - MH: Manhours (auto-calculated as DAYS × 8 if left blank)", False),
            ("", False),
            ("3. Each scope must have at least one task", False),
            ("4. Start date must be before or equal to end date", False),
            ("5. Do not skip rows between tasks", False),
            ("6. Weekly columns (WK1, WK2, etc.) are for visualization only", False),
            ("", False),
            ("DATE FORMAT EXAMPLES:", True),
            ("   Correct: 02-Aug-24, 15-Sep-24, 01-Jan-25", False),
            ("   Incorrect: 8/2/24, 2024-08-02, Aug 2 2024", False),
            ("", False),
            ("IMPORTANT NOTES:", True),
            ("   - Save the file before uploading to the system", False),
            ("   - Maximum 5 upload attempts allowed", False),
            ("   - Schedule requires OM or EG approval before tasks are created", False),
            ("   - Once approved, the schedule cannot be modified", False),
        ]

        row = 1
        for text, is_bold in instructions:
            cell = ws.cell(row=row, column=1, value=text)
            if is_bold:
                cell.font = Font(bold=True, size=12)
            else:
                cell.font = Font(size=11)
            row += 1

        # Set column width
        ws.column_dimensions['A'].width = 80

    def _create_schedule_sheet(self):
        """Create main schedule sheet with project info and scope sections"""
        ws = self.wb.create_sheet("Project Schedule", 1)

        current_row = 1

        # Add project header
        current_row = self._add_project_header(ws, current_row)
        current_row += 1  # Blank row

        # Add scope sections
        for scope in self.scopes:
            current_row = self._add_scope_section(ws, scope, current_row)
            current_row += 1  # Blank row between scopes

        # Freeze panes at header row
        ws.freeze_panes = 'A6'

    def _add_project_header(self, ws, start_row):
        """Add project information header"""
        # Row 1: Project name and code
        ws.merge_cells(f'A{start_row}:E{start_row}')
        cell = ws.cell(row=start_row, column=1,
                      value=f"PROJECT: {self.project.project_name}")
        cell.font = Font(bold=True, size=14)
        cell.fill = PatternFill(start_color=self.PROJECT_INFO_COLOR,
                               end_color=self.PROJECT_INFO_COLOR,
                               fill_type="solid")

        ws.merge_cells(f'F{start_row}:G{start_row}')
        cell = ws.cell(row=start_row, column=6,
                      value=f"Code: {self.project.project_id}")
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color=self.PROJECT_INFO_COLOR,
                               end_color=self.PROJECT_INFO_COLOR,
                               fill_type="solid")

        return start_row

    def _add_scope_section(self, ws, scope, start_row):
        """Add a scope section with headers and empty task rows"""
        # Scope header
        ws.merge_cells(f'A{start_row}:G{start_row}')
        cell = ws.cell(row=start_row, column=1,
                      value=f"SCOPE: {scope.name}")
        cell.font = Font(bold=True, size=13, color="FFFFFF")
        cell.fill = PatternFill(start_color=self.SCOPE_COLOR,
                               end_color=self.SCOPE_COLOR,
                               fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center')
        start_row += 1

        # "PROJECT SCHEDULE" label
        ws.merge_cells(f'A{start_row}:G{start_row}')
        cell = ws.cell(row=start_row, column=1, value="PROJECT SCHEDULE")
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center')
        start_row += 1

        # Column headers
        headers = ['ITEM', 'ACTIVITY', 'START', 'END', 'DAYS', 'MH']

        # Add weekly columns (example: 10 weeks)
        for i in range(1, 11):
            headers.append(f'WK {i}')

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.HEADER_COLOR,
                                   end_color=self.HEADER_COLOR,
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        start_row += 1

        # Add 15 empty rows for task entry
        for i in range(15):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=start_row, column=col_idx, value="")
                cell.border = thin_border

                # Center align specific columns
                if col_idx in [1, 5, 6]:  # ITEM, DAYS, MH
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            start_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 8   # ITEM
        ws.column_dimensions['B'].width = 40  # ACTIVITY
        ws.column_dimensions['C'].width = 12  # START
        ws.column_dimensions['D'].width = 12  # END
        ws.column_dimensions['E'].width = 8   # DAYS
        ws.column_dimensions['F'].width = 8   # MH

        # Weekly columns
        for i in range(7, len(headers) + 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 5

        return start_row

    def _get_filename(self):
        """Generate filename for template"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{self.project.project_id}_Schedule_Template_{timestamp}.xlsx"


def generate_schedule_template(project):
    """
    Main function to generate schedule template for a project

    Args:
        project: ProjectProfile instance

    Returns:
        str: Relative path to generated file (for FileField)
    """
    generator = ScheduleTemplateGenerator(project)
    full_path = generator.generate()

    # Return relative path for Django FileField
    relative_path = os.path.relpath(full_path, settings.MEDIA_ROOT)
    return relative_path
