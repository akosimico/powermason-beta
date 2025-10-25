"""
Excel Schedule File Reader
Parses uploaded schedule files and validates against project scopes
"""

import os
from datetime import datetime, date
from decimal import Decimal
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import logging

logger = logging.getLogger(__name__)


class ScheduleReader:
    """Parse and validate uploaded schedule Excel files"""

    def __init__(self, file_path, project):
        """
        Initialize reader with file and project

        Args:
            file_path: Path to uploaded Excel file
            project: ProjectProfile instance
        """
        self.file_path = file_path
        self.project = project
        self.scopes = {scope.name.strip().upper(): scope
                      for scope in project.scopes.filter(is_deleted=False)}
        self.errors = []
        self.warnings = []
        self.parsed_data = []

    def parse(self):
        """
        Main parsing function

        Returns:
            dict: Parsed data with validation results
        """
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
        except InvalidFileException as e:
            return {
                'success': False,
                'errors': [f"Invalid Excel file format: {str(e)}"],
                'warnings': [],
                'task_count': 0,
                'scopes': []
            }
        except Exception as e:
            return {
                'success': False,
                'errors': [f"Error reading file: {str(e)}"],
                'warnings': [],
                'task_count': 0,
                'scopes': []
            }

        # Look for "Project Schedule" sheet
        if "Project Schedule" in wb.sheetnames:
            ws = wb["Project Schedule"]
        else:
            return {
                'success': False,
                'errors': ['Sheet "Project Schedule" not found in the Excel file'],
                'warnings': [],
                'task_count': 0,
                'scopes': []
            }

        self._parse_schedule_sheet(ws)

        total_tasks = sum(len(scope_data['tasks']) for scope_data in self.parsed_data)

        # Convert dates to ISO format strings for JSON serialization
        serialized_data = self._serialize_dates(self.parsed_data)

        return {
            'success': len(self.errors) == 0,
            'errors': self.errors,
            'warnings': self.warnings,
            'task_count': total_tasks,
            'scopes': serialized_data
        }

    def _serialize_dates(self, data):
        """Convert date objects to ISO format strings for JSON serialization"""
        import copy
        serialized = copy.deepcopy(data)

        for scope in serialized:
            for task in scope.get('tasks', []):
                if task.get('start_date') and isinstance(task['start_date'], date):
                    task['start_date'] = task['start_date'].isoformat()
                if task.get('end_date') and isinstance(task['end_date'], date):
                    task['end_date'] = task['end_date'].isoformat()
                # Convert Decimal to float for JSON
                if task.get('duration_days'):
                    task['duration_days'] = float(task['duration_days'])
                if task.get('manhours'):
                    task['manhours'] = float(task['manhours'])

        return serialized

    def _parse_schedule_sheet(self, ws):
        """Parse the schedule sheet and extract scope sections"""
        max_row = ws.max_row
        current_row = 1

        # Find project header (skip it)
        while current_row <= max_row:
            cell_value = ws.cell(row=current_row, column=1).value
            if cell_value and str(cell_value).upper().startswith("PROJECT:"):
                current_row += 1
                break
            current_row += 1

        # Process scope sections
        while current_row <= max_row:
            cell_value = ws.cell(row=current_row, column=1).value

            if cell_value and str(cell_value).upper().startswith("SCOPE:"):
                # Found a scope section
                scope_name = str(cell_value).replace("SCOPE:", "").strip()
                current_row, scope_data = self._parse_scope_section(ws, current_row, scope_name)

                if scope_data:
                    self.parsed_data.append(scope_data)
            else:
                current_row += 1

    def _parse_scope_section(self, ws, start_row, scope_name):
        """
        Parse a single scope section

        Returns:
            tuple: (next_row, scope_data_dict)
        """
        # Validate scope name against project scopes
        scope_name_upper = scope_name.strip().upper()
        matched_scope = self.scopes.get(scope_name_upper)

        if not matched_scope:
            self.errors.append(
                f"Scope '{scope_name}' not found in project. "
                f"Available scopes: {', '.join(self.scopes.keys())}"
            )
            # Skip this section but continue parsing
            return self._skip_to_next_scope(ws, start_row + 1), None

        scope_data = {
            'scope_name': scope_name,
            'scope_id': matched_scope.id,
            'scope_weight': float(matched_scope.weight),
            'tasks': []
        }

        # Skip "SCOPE:" row
        current_row = start_row + 1

        # Skip "PROJECT SCHEDULE" row if present
        cell_value = ws.cell(row=current_row, column=1).value
        if cell_value and "PROJECT SCHEDULE" in str(cell_value).upper():
            current_row += 1

        # Find header row (ITEM, ACTIVITY, START, END, DAYS, MH)
        header_row = current_row
        headers = []
        for col in range(1, 20):  # Check first 20 columns
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip().upper())
            else:
                headers.append(None)

        # Validate required headers
        required_headers = ['ITEM', 'ACTIVITY', 'START', 'END', 'DAYS', 'MH']
        header_indices = {}

        for req_header in required_headers:
            try:
                header_indices[req_header] = headers.index(req_header)
            except ValueError:
                self.errors.append(
                    f"Missing required column '{req_header}' in scope '{scope_name}'"
                )
                return self._skip_to_next_scope(ws, header_row + 1), None

        # Parse task rows
        current_row = header_row + 1
        task_number = 1

        while current_row <= ws.max_row:
            # Check if we've hit the next scope or end of data
            first_cell = ws.cell(row=current_row, column=1).value
            if first_cell and str(first_cell).upper().startswith("SCOPE:"):
                break

            # Parse task data
            task_data = self._parse_task_row(
                ws, current_row, header_indices, scope_name, task_number
            )

            if task_data:
                scope_data['tasks'].append(task_data)
                task_number += 1
            elif first_cell is None:
                # Empty row - check if all cells are empty
                all_empty = True
                for col in range(1, 7):
                    if ws.cell(row=current_row, column=col).value is not None:
                        all_empty = False
                        break
                if all_empty:
                    # Assume end of scope section if multiple empty rows
                    break

            current_row += 1

        # Validate scope has at least one task
        if len(scope_data['tasks']) == 0:
            self.warnings.append(f"Scope '{scope_name}' has no tasks defined")

        return current_row, scope_data

    def _parse_task_row(self, ws, row, header_indices, scope_name, task_number):
        """
        Parse a single task row

        Returns:
            dict: Task data or None if row is empty/invalid
        """
        # Extract cell values
        item_num = ws.cell(row=row, column=header_indices['ITEM'] + 1).value
        activity = ws.cell(row=row, column=header_indices['ACTIVITY'] + 1).value
        start_date = ws.cell(row=row, column=header_indices['START'] + 1).value
        end_date = ws.cell(row=row, column=header_indices['END'] + 1).value
        days = ws.cell(row=row, column=header_indices['DAYS'] + 1).value
        manhours = ws.cell(row=row, column=header_indices['MH'] + 1).value

        # Skip empty rows (no activity name)
        if not activity:
            return None

        task_data = {
            'item_number': str(item_num) if item_num else f"{task_number}.0",
            'task_name': str(activity).strip(),
            'start_date': None,
            'end_date': None,
            'duration_days': None,
            'manhours': None
        }

        # Parse dates
        task_data['start_date'] = self._parse_date(start_date, row, 'START', scope_name)
        task_data['end_date'] = self._parse_date(end_date, row, 'END', scope_name)

        # Validate date logic
        if task_data['start_date'] and task_data['end_date']:
            if task_data['start_date'] > task_data['end_date']:
                self.errors.append(
                    f"Row {row}: Task '{activity}' has START date after END date"
                )
                return None

            # Calculate duration if not provided
            if not days:
                calculated_days = (task_data['end_date'] - task_data['start_date']).days + 1
                task_data['duration_days'] = Decimal(str(calculated_days))
            else:
                task_data['duration_days'] = self._parse_decimal(days, row, 'DAYS')

            # Calculate manhours if not provided
            if not manhours and task_data['duration_days']:
                task_data['manhours'] = task_data['duration_days'] * 8
            else:
                task_data['manhours'] = self._parse_decimal(manhours, row, 'MH')

        elif not task_data['start_date']:
            self.errors.append(f"Row {row}: Task '{activity}' missing START date")
            return None
        elif not task_data['end_date']:
            self.errors.append(f"Row {row}: Task '{activity}' missing END date")
            return None

        return task_data

    def _parse_date(self, value, row, col_name, scope_name):
        """Parse date value from Excel cell"""
        if value is None:
            return None

        # If already a date object
        if isinstance(value, date):
            return value

        # If datetime object
        if isinstance(value, datetime):
            return value.date()

        # Try to parse string
        if isinstance(value, str):
            # Common formats: "02-Aug-24", "2-Aug-24", "02/08/24", "2024-08-02"
            for fmt in ["%d-%b-%y", "%d-%b-%Y", "%d/%m/%y", "%Y-%m-%d", "%m/%d/%Y"]:
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue

            self.errors.append(
                f"Row {row}: Invalid date format in {col_name} column. "
                f"Use format: DD-MMM-YY (e.g., 02-Aug-24)"
            )
            return None

        self.errors.append(f"Row {row}: Unrecognized date value in {col_name} column")
        return None

    def _parse_decimal(self, value, row, col_name):
        """Parse decimal value from Excel cell"""
        if value is None:
            return Decimal('0')

        try:
            return Decimal(str(value))
        except:
            self.warnings.append(
                f"Row {row}: Invalid number in {col_name} column, using 0"
            )
            return Decimal('0')

    def _skip_to_next_scope(self, ws, start_row):
        """Skip to the next SCOPE: row or end of sheet"""
        current_row = start_row
        while current_row <= ws.max_row:
            cell_value = ws.cell(row=current_row, column=1).value
            if cell_value and str(cell_value).upper().startswith("SCOPE:"):
                return current_row
            current_row += 1
        return current_row


def parse_schedule_excel(file_path, project):
    """
    Main function to parse uploaded schedule Excel file

    Args:
        file_path: Path to Excel file
        project: ProjectProfile instance

    Returns:
        dict: Parsing results with success flag, errors, and data
    """
    reader = ScheduleReader(file_path, project)
    return reader.parse()
